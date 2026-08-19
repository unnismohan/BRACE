"""
BRACE v2 — FastAPI controller
Production-ready test management platform for Robot Framework.
"""

import asyncio
import io
import logging
import os
import re
import shutil
import subprocess
import time
import uuid
import xml.etree.ElementTree as ET
import zipfile
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import Depends, FastAPI, File, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, PlainTextResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from fastapi.staticfiles import StaticFiles
from jose import JWTError, jwt
from pydantic import BaseModel

from db import (
    DB_PATH, decrypt_token, encrypt_token, get_db, init_db,
    next_tc_code, pwd_context, row_to_dict, rows_to_list,
)
from scheduler import (
    SCHEDULER_TZ, build_trigger, next_run_times, reload_schedules,
    scheduled_job_ids, start_scheduler, stop_scheduler,
)
import git_sync
import mailer
import maintenance

log = logging.getLogger(__name__)


class _JsonLogFormatter(logging.Formatter):
    """One JSON object per line, so a log stack can index the fields.

    Anything passed via `extra=` (run_id, project_id, user…) is merged in, which
    is what makes a run traceable across the many log lines it produces.
    """
    _BUILTIN = set(logging.LogRecord("", 0, "", 0, "", (), None).__dict__) | {
        "asctime", "message", "taskName"}

    def format(self, record: logging.LogRecord) -> str:
        import json as _json
        out = {
            "ts":      datetime.fromtimestamp(record.created).isoformat(timespec="milliseconds"),
            "level":   record.levelname,
            "logger":  record.name,
            "message": record.getMessage(),
        }
        for k, v in record.__dict__.items():
            if k not in self._BUILTIN and not k.startswith("_"):
                out[k] = v
        if record.exc_info:
            out["exception"] = self.formatException(record.exc_info)
        return _json.dumps(out, default=str)


def _configure_logging() -> None:
    fmt = os.getenv("BRACE_LOG_FORMAT", "text").strip().lower()
    handler = logging.StreamHandler()
    if fmt == "json":
        handler.setFormatter(_JsonLogFormatter())
    else:
        handler.setFormatter(logging.Formatter(
            "%(asctime)s %(levelname)s %(name)s: %(message)s"))
    root = logging.getLogger()
    root.handlers[:] = [handler]
    root.setLevel(os.getenv("BRACE_LOG_LEVEL", "INFO").strip().upper() or "INFO")


_configure_logging()

# ── Config ────────────────────────────────────────────────────────
SUITES_DIR  = Path(os.getenv("SUITES_DIR",  "/opt/rf/suites"))
RESULTS_DIR = Path(os.getenv("RESULTS_DIR", "/opt/rf/results"))
# Application version. Single source of truth — the Dockerfile passes it in as
# BRACE_VERSION from its APP_VERSION build arg, which also stamps the image
# label, so /health and the image can never disagree. The literal below is only
# the fallback for running outside the container.
APP_VERSION = os.getenv("BRACE_VERSION", "").strip() or "2.2.0"

BSS_ENV     = os.getenv("BSS_ENV",    "staging")
IMAGE_TAG   = os.getenv("IMAGE_TAG",  "unknown")
POD_NAME    = os.getenv("HOSTNAME",   "unknown")

JWT_SECRET     = os.getenv("JWT_SECRET", "brace-default-secret-change-in-prod")
JWT_ALGORITHM  = "HS256"
JWT_EXPIRE_MIN = 480  # 8 h
SSE_TIMEOUT    = 1800  # 30 min

ALLOWED_EXTS = {".robot", ".yaml", ".yml", ".txt", ".py", ".resource", ".csv", ".xlsx"}

# ── Execution capacity ────────────────────────────────────────────
# Two separate limits, because they bound different things:
#
#   MAX_CONCURRENT_RUNS  — how many runs are admitted at once. Purely a
#                          queueing/fairness knob; extra runs wait.
#   MAX_CONCURRENT_TESTS — how many robot processes (each = one Chrome +
#                          chromedriver) execute at any moment across every
#                          run. This is the one bounded by physical resources.
#
# Test cases inside a run execute in parallel, so a single run can occupy the
# whole browser budget. Without that second cap, one 44-case run would spawn 44
# browsers and the pod would be OOM-killed, losing every in-flight run.
#
# Sizing guide (pod resources -> sensible MAX_CONCURRENT_TESTS):
#     2 CPU /  4Gi  -> 3   (current deployment)
#     4 CPU /  8Gi  -> 6
#     8 CPU / 16Gi  -> 12
# Raise it in the Deployment only together with the pod's CPU/memory limits.
MAX_CONCURRENT_RUNS = max(1, int(os.getenv("BRACE_MAX_CONCURRENT_RUNS", "3")))

# The real physical limit is browsers, not runs: test cases inside one run now
# execute in parallel, so "3 runs" no longer means "3 Chromes". This is the
# total robot processes in flight across every run — the number that must match
# the pod's CPU/memory. It defaults to MAX_CONCURRENT_RUNS, which reproduces the
# old resource ceiling exactly; raise it only when you raise the pod's limits.
MAX_CONCURRENT_TESTS = max(1, int(os.getenv("BRACE_MAX_CONCURRENT_TESTS",
                                            str(MAX_CONCURRENT_RUNS))))

# How many test cases one run may execute at once, capped by the global budget
# above. Default: the whole budget, so a single run finishes as fast as the pod
# allows instead of trickling through one case at a time. Set to 1 for suites
# whose cases must run in order — or pass "parallel" on the run request.
RUN_PARALLEL_DEFAULT = max(1, min(
    MAX_CONCURRENT_TESTS,
    int(os.getenv("BRACE_RUN_PARALLEL", str(MAX_CONCURRENT_TESTS)))))

# Hard ceiling per test case. A hung browser would otherwise occupy a slot
# forever and block the whole queue.
TEST_TIMEOUT_SEC = max(60, int(os.getenv("BRACE_TEST_TIMEOUT_SEC", "1800")))  # 30 min

# In-memory run state
_active_runs: dict = {}   # run_id → status dict
_active_procs: dict = {}  # rf_run_id → asyncio Process
_cancelled_runs: set = set()   # run_ids cancelled before their slot came up
_run_slots: Optional[asyncio.Semaphore] = None   # created in lifespan (needs a loop)
_test_slots: Optional[asyncio.Semaphore] = None  # global robot-process budget
_main_loop: Optional[asyncio.AbstractEventLoop] = None  # for scheduler thread hand-off

# ── Live run events (SSE) ────────────────────────────────────────
# run_id → set of subscriber queues. In-process is the right scope: runs execute
# in this pod and the Deployment is strategy: Recreate with one replica, so
# there is no second process that could hold a subscriber for a run it is not
# running. A multi-replica deployment would need a broker here instead.
_run_subs: dict = {}
# Per-run subscriber ceiling. Ten people watching one run is already unusual;
# past that the browsers fall back to polling rather than the pod fanning out
# indefinitely.
SSE_MAX_SUBS = max(1, int(os.getenv("BRACE_SSE_MAX_SUBS", "20")))
SSE_HEARTBEAT_SEC = 15    # keeps the OCP router from idling the connection out


def _publish(run_id: str, event: str, data: dict) -> None:
    """Fan one event out to everyone watching this run. Never raises, never blocks.

    Called from the executor's hot path, so a slow or dead subscriber must not
    be able to stall a test finishing: queues are unbounded-put via put_nowait
    and a full one simply drops the event — the client's next reconnect (or the
    fallback poll) resynchronises from the database, which is the source of
    truth either way.
    """
    subs = _run_subs.get(run_id)
    if not subs:
        return
    for q in list(subs):
        try:
            q.put_nowait((event, data))
        except Exception:                              # noqa: BLE001 — see docstring
            pass

# Process-lifetime counters for /metrics. Reset on restart, which is normal for
# Prometheus counters — it detects the reset and handles it.
_metrics = {
    "runs_started":   0,
    "runs_passed":    0,
    "runs_failed":    0,
    "runs_cancelled": 0,
    "tests_passed":   0,
    "tests_failed":   0,
    "tests_timeout":  0,
    "test_seconds":   0.0,
    "test_count":     0,
    "started_at":     datetime.now(),
}


def _slots() -> asyncio.Semaphore:
    """Lazily create the semaphore — it must bind to the running event loop."""
    global _run_slots
    if _run_slots is None:
        _run_slots = asyncio.Semaphore(MAX_CONCURRENT_RUNS)
    return _run_slots


def _tslots() -> asyncio.Semaphore:
    """Global cap on concurrent robot processes (i.e. concurrent browsers)."""
    global _test_slots
    if _test_slots is None:
        _test_slots = asyncio.Semaphore(MAX_CONCURRENT_TESTS)
    return _test_slots


# ── Lifespan ─────────────────────────────────────────────────────
def _preflight_security_check() -> None:
    """Refuse to start a non-local deployment with insecure defaults."""
    problems, warnings = [], []

    if JWT_SECRET == "brace-default-secret-change-in-prod":
        problems.append("JWT_SECRET is the built-in default — anyone can forge a login token. "
                        "Set it to a long random value.")
    elif len(JWT_SECRET) < 32:
        warnings.append(f"JWT_SECRET is only {len(JWT_SECRET)} chars; use 32+.")

    if not os.getenv("BRACE_ENCRYPT_KEY", "").strip():
        warnings.append("BRACE_ENCRYPT_KEY is not set — git tokens, the AI API key and "
                        "the SMTP password are stored in plain text in the database.")

    # A container with no TZ runs in UTC while the team reads the UI in their own
    # zone, so every timestamp looks hours out. Cheap to detect, confusing to
    # diagnose from the symptom.
    tz_env = (os.getenv("TZ") or "").strip()
    if not tz_env:
        warnings.append(
            f"TZ is not set, so the container clock is {time.tzname[0]}. Timestamps are "
            f"stored and displayed in that zone — set TZ (usually to the same value as "
            f"BRACE_SCHEDULER_TZ={SCHEDULER_TZ}) or times in the UI will not match "
            f"anyone's watch.")
    elif tz_env != SCHEDULER_TZ:
        warnings.append(
            f"TZ ({tz_env}) and BRACE_SCHEDULER_TZ ({SCHEDULER_TZ}) differ. Cron "
            f"expressions will fire on one clock while timestamps are recorded on "
            f"another — set both the same unless you specifically want this.")

    try:
        conn = get_db()
        row = conn.execute(
            "SELECT must_change_password FROM users WHERE username='admin'").fetchone()
        conn.close()
        if row and row["must_change_password"]:
            warnings.append("The bootstrap 'admin' account still has its initial password. "
                            "Change it immediately after first login.")
    except Exception:                              # noqa: BLE001 — never block startup on this probe
        pass

    for w in warnings:
        log.warning("SECURITY: %s", w)

    if problems:
        for p in problems:
            log.error("SECURITY: %s", p)
        if BSS_ENV.lower() not in ("local", "dev", "development"):
            raise RuntimeError(
                "Refusing to start with insecure defaults: " + " | ".join(problems)
                + "  (set BSS_ENV=local to bypass for local development)")
        log.warning("SECURITY: continuing anyway because BSS_ENV=%s", BSS_ENV)


def _reconcile_orphaned_runs() -> None:
    """Close out runs left mid-flight by a previous pod.

    Run state lives in this process (subprocesses + in-memory dicts), so
    anything still marked queued/running in the DB at startup belongs to a pod
    that no longer exists. Without this they linger forever and the UI keeps
    polling them as if they were live.
    """
    conn = get_db()
    now = datetime.now().isoformat()
    items = conn.execute(
        "UPDATE test_run_items SET status='cancelled', finished_at=?"
        " WHERE status IN ('running','pending','queued')", (now,)).rowcount
    runs = conn.execute(
        "UPDATE test_runs SET status='cancelled', finished_at=?"
        " WHERE status IN ('running','queued')", (now,)).rowcount
    conn.commit()
    conn.close()
    if runs or items:
        log.warning("Startup: cancelled %d orphaned run(s) and %d test item(s) "
                    "left behind by a previous pod.", runs, items)


def _audit_route_guards() -> None:
    """Warn at startup about any mutating API route with no authentication.

    Cheap insurance for the thing that goes wrong when a codebase grows a fourth
    role: one new @app.post added without a Depends, silently open to anyone.
    A test would catch it only if someone remembered to write one; this catches
    it every boot.
    """
    guards = {"_current_user", "_require_sys_admin", "dep"}     # 'dep' = _require_project_role
    # Routes that are public by design. Anything not on this list must be
    # authenticated, so a new endpoint added without a Depends is reported.
    public = {"/api/auth/login"}
    unguarded = []
    for route in app.routes:
        methods = getattr(route, "methods", set()) or set()
        path    = getattr(route, "path", "")
        if (path in public or not path.startswith("/api/")
                or not (methods & {"POST", "PUT", "PATCH", "DELETE"})):
            continue
        dependant = getattr(route, "dependant", None)
        if dependant is None:
            continue
        found, stack = False, list(dependant.dependencies)
        while stack and not found:
            d = stack.pop()
            name = getattr(d.call, "__name__", "")
            if name in guards or "oauth2" in name.lower():
                found = True
            stack.extend(d.dependencies)
        if not found:
            unguarded.append(f"{'/'.join(sorted(methods & {'POST','PUT','PATCH','DELETE'}))} {path}")
    if unguarded:
        log.error("SECURITY: %d mutating API route(s) have no auth dependency: %s",
                  len(unguarded), ", ".join(sorted(unguarded)))


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    _preflight_security_check()
    _audit_route_guards()
    _reconcile_orphaned_runs()
    _slots()                       # bind the semaphore to this event loop
    global _main_loop
    _main_loop = asyncio.get_running_loop()   # scheduler threads marshal onto this
    start_scheduler()
    _reload_all_jobs()
    # Log the wall clock the container actually has. Timestamps are stored as
    # local time and rendered as-is, so if this line does not match the reader's
    # watch, every time in the UI will look wrong — and it is a one-line check.
    _now = datetime.now()
    log.info("BRACE v2 started — env=%s tag=%s version=%s max_concurrent_runs=%d "
             "max_concurrent_tests=%d run_parallel=%d test_timeout=%ds "
             "scheduler_tz=%s container_tz=%s local_time=%s retention=%s",
             BSS_ENV, IMAGE_TAG, APP_VERSION, MAX_CONCURRENT_RUNS, MAX_CONCURRENT_TESTS,
             RUN_PARALLEL_DEFAULT, TEST_TIMEOUT_SEC, SCHEDULER_TZ,
             os.getenv("TZ") or time.tzname[0], _now.strftime("%Y-%m-%d %H:%M:%S"),
             f"{maintenance.RETENTION_DAYS}d (keep min {maintenance.RETENTION_KEEP_MIN}/project)"
             if maintenance.RETENTION_DAYS else "off")
    yield
    stop_scheduler()


app = FastAPI(title="BRACE RF Controller", version=APP_VERSION, lifespan=lifespan)

# ── Auth ──────────────────────────────────────────────────────────
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)


def _make_token(username: str, system_role: str) -> str:
    exp = datetime.utcnow() + timedelta(minutes=JWT_EXPIRE_MIN)
    return jwt.encode({"sub": username, "role": system_role, "exp": exp}, JWT_SECRET, algorithm=JWT_ALGORITHM)


def _current_user(token: str = Depends(oauth2_scheme)) -> dict:
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return {"username": payload["sub"], "role": payload.get("role", "user")}
    except JWTError:
        raise HTTPException(401, "Invalid or expired token")


def _require_sys_admin(user=Depends(_current_user)):
    if user["role"] != "admin":
        raise HTTPException(403, "System admin required")
    return user


def _project_suites(project_id: int) -> Path:
    return SUITES_DIR / str(project_id)


def _project_results(project_id: int) -> Path:
    p = RESULTS_DIR / str(project_id)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _get_project_role(project_id: int, username: str, user_system_role: str) -> Optional[str]:
    """Return effective project role. Sys admins have project_admin everywhere."""
    if user_system_role == "admin":
        return "project_admin"
    conn = get_db()
    row = conn.execute(
        """SELECT pm.project_role FROM project_members pm
           JOIN users u ON pm.user_id = u.id
           WHERE pm.project_id=? AND u.username=?""",
        (project_id, username),
    ).fetchone()
    conn.close()
    return row["project_role"] if row else None


def _require_project_role(*roles: str):
    """Dependency factory — requires one of the given project roles."""
    def dep(project_id: int, user=Depends(_current_user)):
        role = _get_project_role(project_id, user["username"], user["role"])
        if role not in roles:
            raise HTTPException(403, f"Requires project role: {', '.join(roles)}")
        return {**user, "project_role": role}
    return dep


_proj_viewer  = _require_project_role("viewer", "tester", "project_admin")
_proj_tester  = _require_project_role("tester", "project_admin")
_proj_admin   = _require_project_role("project_admin")

# What each project role may do. The three dependencies above cover endpoints
# whose path carries {project_id}; this map is the same rule expressed for the
# endpoints that resolve the project from a resource id (a test case, a suite, a
# run) and so cannot use a path-parameter dependency.
ROLE_CAPS = {
    "viewer":        {"view"},
    "tester":        {"view", "run", "edit"},
    "project_admin": {"view", "run", "edit", "manage"},
}


def _require_cap(project_id: int, user: dict, cap: str) -> str:
    """Authorise `user` for `cap` on `project_id`, or raise 403.

    Used by resource-scoped endpoints. Every one of these previously carried its
    own inline role tuple, and they had already drifted apart — deleting a suite
    demanded project_admin in one place and accepted tester in another.
    """
    role = _get_project_role(project_id, user["username"], user["role"])
    if cap not in ROLE_CAPS.get(role or "", ()):
        raise HTTPException(403, f"Requires '{cap}' permission on this project"
                                 f" (your role: {role or 'none'})")
    return role


def _owned_row(conn, sql: str, ident, user: dict, cap: str, missing: str):
    """Fetch a resource row, authorise the caller against its project, or raise.

    Closes the connection on both failure paths — the callers open it before
    they can know which project the resource belongs to, and an early `raise`
    without this leaks the handle.
    """
    row = conn.execute(sql, (ident,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, missing)
    try:
        _require_cap(row["project_id"], user, cap)
    except HTTPException:
        conn.close()
        raise
    return row


# ── Audit trail ──────────────────────────────────────────────────
# Records who changed what. Three rules, all of them load-bearing:
#   1. It never raises. An audit write failing must not fail the operation the
#      user asked for — a lost audit line is bad, a lost test run is worse.
#   2. It records mutations only. Logging reads would bury the interesting lines
#      under thousands of page views.
#   3. It never records a secret. Passwords, tokens and API keys are recorded as
#      the fact that they changed, never the value.
_AUDIT_SECRET_KEYS = ("password", "token", "api_key", "secret", "key")


def audit(user, action: str, project_id: Optional[int] = None,
          target: Optional[str] = None, **detail) -> None:
    import json as _json
    try:
        username = user.get("username") if isinstance(user, dict) else str(user or "")
        safe = {}
        for k, v in detail.items():
            if any(s in k.lower() for s in _AUDIT_SECRET_KEYS):
                # Keep the signal (it changed), drop the value.
                safe[k] = bool(v) if not isinstance(v, bool) else v
            else:
                safe[k] = v
        conn = get_db()
        try:
            conn.execute(
                "INSERT INTO audit_log (ts, username, action, project_id, target, detail)"
                " VALUES (?,?,?,?,?,?)",
                (datetime.now().isoformat(timespec="seconds"), username, action,
                 project_id, None if target is None else str(target),
                 _json.dumps(safe, default=str) if safe else None))
            conn.commit()
        finally:
            conn.close()
    except Exception as exc:                       # noqa: BLE001 — see rule 1
        log.debug("Audit write failed for %s: %s", action, exc)


# ══════════════════════════════════════════════════════════════════
# AUTH ENDPOINTS
# ══════════════════════════════════════════════════════════════════

class LoginRequest(BaseModel):
    username: str
    password: str


class ChangePwRequest(BaseModel):
    old_password: str
    new_password: str


@app.post("/api/auth/login")
def login(req: LoginRequest):
    conn = get_db()
    row = conn.execute("SELECT * FROM users WHERE username=?", (req.username,)).fetchone()
    conn.close()
    if not row or not pwd_context.verify(req.password, row["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    token = _make_token(row["username"], row["system_role"])
    return {
        "access_token":        token,
        "token_type":          "bearer",
        "username":            row["username"],
        "system_role":         row["system_role"],
        "must_change_password": bool(row["must_change_password"]),
    }


@app.put("/api/auth/change-password")
def change_password(req: ChangePwRequest, user=Depends(_current_user)):
    conn = get_db()
    row = conn.execute("SELECT * FROM users WHERE username=?", (user["username"],)).fetchone()
    if not row or not pwd_context.verify(req.old_password, row["password_hash"]):
        conn.close()
        raise HTTPException(400, "Current password incorrect")
    conn.execute(
        "UPDATE users SET password_hash=?, must_change_password=0 WHERE username=?",
        (pwd_context.hash(req.new_password), user["username"]),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════
# USER MANAGEMENT (sys admin)
# ══════════════════════════════════════════════════════════════════

class UserCreate(BaseModel):
    username:    str
    password:    str
    system_role: str = "user"
    full_name:   Optional[str] = None
    email:       Optional[str] = None


class UserUpdate(BaseModel):
    system_role: Optional[str] = None
    password:    Optional[str] = None
    full_name:   Optional[str] = None
    email:       Optional[str] = None


@app.get("/api/users", dependencies=[Depends(_require_sys_admin)])
def list_users():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, username, system_role, full_name, email, must_change_password, created_at FROM users ORDER BY username"
    ).fetchall()
    conn.close()
    return rows_to_list(rows)


@app.post("/api/users")
def create_user(req: UserCreate, user=Depends(_require_sys_admin)):
    if req.system_role not in ("user", "admin"):
        raise HTTPException(400, "system_role must be user or admin")
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, system_role, full_name, email) VALUES (?,?,?,?,?)",
            (req.username, pwd_context.hash(req.password), req.system_role, req.full_name, req.email),
        )
        conn.commit()
    except Exception:
        conn.close()
        raise HTTPException(409, "Username already exists")
    uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    audit(user, "user.create", target=req.username, system_role=req.system_role)
    return {"id": uid, "username": req.username, "system_role": req.system_role}


@app.post("/api/users/bulk-csv")
async def bulk_create_users(file: UploadFile = File(...), user=Depends(_require_sys_admin)):
    """CSV: username, password, system_role (opt), full_name (opt), email (opt)"""
    import csv
    content = (await file.read()).decode("utf-8-sig")
    reader  = csv.DictReader(io.StringIO(content))
    created, skipped = [], []
    conn = get_db()
    for row in reader:
        username = (row.get("username") or "").strip()
        password = (row.get("password") or "").strip()
        role     = (row.get("system_role") or row.get("role") or "user").strip().lower()
        if not username or not password:
            skipped.append(username or "(empty)")
            continue
        if role not in ("user", "admin"):
            role = "user"
        try:
            conn.execute(
                "INSERT INTO users (username, password_hash, system_role, full_name, email, must_change_password) VALUES (?,?,?,?,?,1)",
                (username, pwd_context.hash(password), role,
                 (row.get("full_name") or "").strip() or None,
                 (row.get("email") or "").strip() or None),
            )
            conn.commit()
            created.append({"username": username, "system_role": role})
        except Exception:
            skipped.append(username)
    conn.close()
    audit(user, "user.bulk_create", created=len(created), skipped=len(skipped))
    return {"created": len(created), "skipped": skipped, "users": created}


@app.put("/api/users/{uid}")
def update_user(uid: int, req: UserUpdate, user=Depends(_require_sys_admin)):
    conn = get_db()
    before = conn.execute("SELECT username, system_role FROM users WHERE id=?",
                          (uid,)).fetchone()
    if req.system_role:
        if req.system_role not in ("user", "admin"):
            conn.close()
            raise HTTPException(400, "Invalid system_role")
        conn.execute("UPDATE users SET system_role=? WHERE id=?", (req.system_role, uid))
    if req.password:
        conn.execute("UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?",
                     (pwd_context.hash(req.password), uid))
    if req.full_name is not None:
        conn.execute("UPDATE users SET full_name=? WHERE id=?", (req.full_name, uid))
    if req.email is not None:
        conn.execute("UPDATE users SET email=? WHERE id=?", (req.email, uid))
    conn.commit()
    conn.close()
    target = before["username"] if before else str(uid)
    # A privilege change is the single most important line in this log, so it
    # gets its own action rather than being buried inside a generic update.
    if req.system_role and before and req.system_role != before["system_role"]:
        audit(user, "user.role_change", target=target,
              **{"from": before["system_role"], "to": req.system_role})
    audit(user, "user.update", target=target,
          password_changed=bool(req.password),
          fields=[f for f in ("full_name", "email") if getattr(req, f) is not None])
    return {"ok": True}


@app.delete("/api/users/{uid}")
def delete_user(uid: int, user=Depends(_require_sys_admin)):
    conn = get_db()
    row = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    audit(user, "user.delete", target=row["username"] if row else str(uid))
    return {"deleted": uid}


# ══════════════════════════════════════════════════════════════════
# PROJECT MANAGEMENT
# ══════════════════════════════════════════════════════════════════

class ProjectCreate(BaseModel):
    name:         str
    description:  Optional[str] = None
    git_url:      Optional[str] = None
    git_branch:   str           = "main"
    git_username: Optional[str] = None
    git_token:    Optional[str] = None


class ProjectUpdate(BaseModel):
    name:        Optional[str] = None
    description: Optional[str] = None
    status:      Optional[str] = None


@app.get("/api/projects")
def list_projects(user=Depends(_current_user)):
    conn = get_db()
    if user["role"] == "admin":
        rows = conn.execute("SELECT * FROM projects ORDER BY name").fetchall()
    else:
        rows = conn.execute("""
            SELECT p.* FROM projects p
            JOIN project_members pm ON p.id = pm.project_id
            JOIN users u ON pm.user_id = u.id
            WHERE u.username=? ORDER BY p.name
        """, (user["username"],)).fetchall()

    result = []
    for p in rows:
        d = dict(p)
        d.pop("git_token", None)  # never expose token
        # Enrich with stats
        stats = conn.execute("""
            SELECT COUNT(*) total,
                   SUM(CASE WHEN last_run_status='passed' THEN 1 ELSE 0 END) passed,
                   SUM(CASE WHEN last_run_status='failed' THEN 1 ELSE 0 END) failed
            FROM test_cases WHERE project_id=?
        """, (d["id"],)).fetchone()
        last_run = conn.execute("""
            SELECT status, started_at FROM test_runs
            WHERE project_id=? ORDER BY started_at DESC, id DESC LIMIT 1
        """, (d["id"],)).fetchone()
        d["tc_count"]     = stats["total"] or 0
        d["tc_passed"]    = stats["passed"] or 0
        d["tc_failed"]    = stats["failed"] or 0
        d["last_run_status"]  = last_run["status"]    if last_run else None
        d["last_run_at"]      = last_run["started_at"] if last_run else None
        d["has_git"]      = bool(d.get("git_url"))
        # Caller's own role on this project — lets the UI hide admin-only views
        d["my_role"] = ("project_admin" if user["role"] == "admin"
                        else (conn.execute(
                            """SELECT pm.project_role FROM project_members pm
                               JOIN users u ON pm.user_id = u.id
                               WHERE pm.project_id=? AND u.username=?""",
                            (d["id"], user["username"])).fetchone() or {"project_role": None})["project_role"])
        result.append(d)
    conn.close()
    return result


@app.post("/api/projects", dependencies=[Depends(_require_sys_admin)])
def create_project(req: ProjectCreate, user=Depends(_current_user)):
    conn = get_db()
    conn.execute(
        "INSERT INTO projects (name, description, git_url, git_branch, git_username, git_token) VALUES (?,?,?,?,?,?)",
        (req.name, req.description, req.git_url, req.git_branch, req.git_username,
         encrypt_token(req.git_token or "")),
    )
    conn.commit()
    pid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    # Creator gets project_admin
    conn.execute("INSERT INTO project_members (project_id, user_id, project_role) VALUES (?,?,?)",
                 (pid, conn.execute("SELECT id FROM users WHERE username=?", (user["username"],)).fetchone()[0],
                  "project_admin"))
    conn.commit()
    row = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    conn.close()
    d = dict(row)
    d.pop("git_token", None)
    audit(user, "project.create", project_id=pid, target=req.name,
          git_url=req.git_url or None)
    return d


@app.get("/api/projects/{project_id}")
def get_project(project_id: int, user=Depends(_proj_viewer)):
    conn = get_db()
    row = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Project not found")
    d = dict(row)
    d.pop("git_token", None)
    d["has_git"] = bool(d.get("git_url"))
    return d


@app.put("/api/projects/{project_id}")
def update_project(project_id: int, req: ProjectUpdate, user=Depends(_proj_admin)):
    conn = get_db()
    if req.name:        conn.execute("UPDATE projects SET name=?        WHERE id=?", (req.name,        project_id))
    if req.description is not None:
                        conn.execute("UPDATE projects SET description=? WHERE id=?", (req.description, project_id))
    if req.status:      conn.execute("UPDATE projects SET status=?      WHERE id=?", (req.status,      project_id))
    conn.commit()
    conn.close()
    audit(user, "project.update", project_id=project_id,
          name=req.name, status=req.status)
    return {"ok": True}


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int, user=Depends(_require_sys_admin)):
    conn = get_db()
    row = conn.execute("SELECT name FROM projects WHERE id=?", (project_id,)).fetchone()
    conn.execute("DELETE FROM projects WHERE id=?", (project_id,))
    conn.commit()
    conn.close()
    audit(user, "project.delete", project_id=project_id,
          target=row["name"] if row else str(project_id))
    return {"deleted": project_id}


# ── Git Config ───────────────────────────────────────────────────
class GitConfigUpdate(BaseModel):
    git_url:      Optional[str] = None
    git_branch:   str           = "main"
    git_username: Optional[str] = None
    git_token:    Optional[str] = None  # None = keep existing


@app.get("/api/projects/{project_id}/git-config")
def get_git_config(project_id: int, user=Depends(_proj_viewer)):
    conn = get_db()
    row = conn.execute("SELECT git_url, git_branch, git_username, git_token FROM projects WHERE id=?",
                       (project_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404)
    return {
        "git_url":      row["git_url"],
        "git_branch":   row["git_branch"],
        "git_username": row["git_username"],
        "has_token":    bool(row["git_token"]),
    }


@app.put("/api/projects/{project_id}/git-config")
def update_git_config(project_id: int, req: GitConfigUpdate, user=Depends(_proj_admin)):
    conn = get_db()
    if req.git_url is not None:
        conn.execute("UPDATE projects SET git_url=? WHERE id=?", (req.git_url, project_id))
    conn.execute("UPDATE projects SET git_branch=? WHERE id=?", (req.git_branch, project_id))
    if req.git_username is not None:
        conn.execute("UPDATE projects SET git_username=? WHERE id=?", (req.git_username, project_id))
    if req.git_token is not None:
        conn.execute("UPDATE projects SET git_token=? WHERE id=?",
                     (encrypt_token(req.git_token), project_id))
    conn.commit()
    conn.close()
    audit(user, "git.config_update", project_id=project_id,
          git_url=req.git_url, git_branch=req.git_branch,
          git_token=req.git_token is not None)   # recorded as True/False, never the value
    return {"ok": True}


@app.post("/api/projects/{project_id}/git-sync")
async def git_pull(project_id: int, user=Depends(_proj_tester)):
    # NB: named git_pull, not git_sync — the module `git_sync` is imported above
    # and a same-named function here would shadow it for the whole file.
    conn = get_db()
    row = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404)

    git_url  = row["git_url"] or ""
    branch   = row["git_branch"] or "main"
    username = row["git_username"] or ""
    token    = decrypt_token(row["git_token"] or "")

    if not git_url:
        raise HTTPException(400, "No git URL configured for this project")

    suites_dir = _project_suites(project_id)
    clone_dir  = Path("/tmp") / f"brace-git-{project_id}"

    if username or token:
        from urllib.parse import urlparse, urlunparse
        p      = urlparse(git_url)
        netloc = f"{username}:{token}@{p.hostname}" if username else f"{token}@{p.hostname}"
        if p.port:
            netloc += f":{p.port}"
        git_url = urlunparse(p._replace(netloc=netloc))

    def _redact(text: str) -> str:
        """git echoes the remote URL on failure — with the embedded PAT in it.
        Strip any credentials before this reaches the browser or the logs."""
        if not text:
            return text
        out = re.sub(r"(https?://)[^/\s@]+@", r"\1***@", text)
        for secret in (token, username):
            if secret and len(secret) > 3:
                out = out.replace(secret, "***")
        return out

    # Off-thread: a git clone can take tens of seconds and this generator runs
    # on the event loop — a blocking call here stalls every other request.
    async def _run(cmd, cwd=None):
        return await asyncio.to_thread(
            subprocess.run, cmd, capture_output=True, text=True, cwd=cwd)

    async def stream():
        yield f"[BRACE] Git sync — project {project_id}, branch: {branch}\n"
        if clone_dir.exists():
            yield "[BRACE] Existing clone — fetching latest…\n"
            # remove any stale git lock files left by crashed processes
            for lock in clone_dir.rglob("*.lock"):
                try:
                    lock.unlink()
                    yield f"[BRACE] Removed stale lock: {lock.name}\n"
                except OSError:
                    pass
            r = await _run(["git", "-c", "http.sslVerify=false", "fetch", "--depth=1", "origin", branch], cwd=clone_dir)
            yield _redact(r.stdout + r.stderr)
            if r.returncode == 0:
                r2 = await _run(["git", "reset", "--hard", f"origin/{branch}"], cwd=clone_dir)
                yield _redact(r2.stdout + r2.stderr)
                if r2.returncode != 0:
                    yield "[BRACE] Reset failed — will re-clone…\n"
                    shutil.rmtree(clone_dir, ignore_errors=True)
            else:
                yield "[BRACE] Fetch failed — will re-clone…\n"
                shutil.rmtree(clone_dir, ignore_errors=True)

        if not clone_dir.exists():
            yield "[BRACE] Cloning repository…\n"
            r = await _run(["git", "-c", "http.sslVerify=false", "clone", "--depth=1", "--branch", branch, git_url, str(clone_dir)])
            yield _redact(r.stdout + r.stderr)
            if r.returncode != 0:
                yield f"\n[BRACE ERROR] Clone failed (exit {r.returncode})\n"
                return

        yield "\n[BRACE] Copying scripts to project suites dir…\n"

        def _copy_all() -> int:
            """Thousands of file copies — run off the event loop."""
            suites_dir.mkdir(parents=True, exist_ok=True)
            n = 0
            for ext in (".robot", ".resource", ".py", ".yaml", ".yml", ".csv", ".xlsx"):
                for s in sorted(clone_dir.rglob(f"*{ext}")):
                    if ".git" in s.parts:
                        continue
                    d = suites_dir / s.relative_to(clone_dir)
                    d.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copyfile(s, d)
                    if ext == ".robot":
                        n += 1
            return n

        copied = await asyncio.to_thread(_copy_all)
        yield f"[BRACE] Done — {copied} robot file(s) synced\n"
        audit(user, "git.pull", project_id=project_id, branch=branch, files=copied)

        # Projects in git mode reconcile their test cases in the same action —
        # pulling scripts and then leaving the case list stale is exactly the
        # drift this feature exists to remove.
        if (row["sync_mode"] or "manual") == "git":
            yield "\n[BRACE] Reconciling test cases from the repository…\n"
            try:
                res = await asyncio.to_thread(_run_tc_sync, project_id, user["username"])
                s = git_sync.summary(res)
                yield (f"[BRACE] Test cases — added {s['added']}, updated {s['updated']}, "
                       f"unchanged {s['unchanged']}, missing {s['missing']}\n")
                for e in res.get("errors", [])[:10]:
                    yield f"[BRACE WARN] {e}\n"
            except Exception as exc:                   # noqa: BLE001 — surfaced, not fatal
                yield f"[BRACE ERROR] Test case sync failed: {exc}\n"

    return StreamingResponse(stream(), media_type="text/plain")


# ── Git-native test cases ────────────────────────────────────────
class SyncConfigReq(BaseModel):
    sync_mode: str                       # 'manual' | 'git'
    sync_cron: Optional[str] = None      # optional automatic reconcile


def _run_tc_sync(project_id: int, username: str, dry_run: bool = False) -> dict:
    """Blocking — parses every .robot file. Always call via to_thread."""
    conn = get_db()
    try:
        res = git_sync.sync_project(conn, project_id, _project_suites(project_id),
                                    next_tc_code, _norm_tags, dry_run=dry_run)
        if not dry_run:
            import json as _json
            conn.execute("UPDATE projects SET last_sync_at=?, last_sync_result=? WHERE id=?",
                         (datetime.now().isoformat(timespec="seconds"),
                          _json.dumps(git_sync.summary(res)), project_id))
            conn.commit()
    finally:
        conn.close()
    if not dry_run:
        s = git_sync.summary(res)
        audit(username, "project.sync", project_id=project_id, **s)
    return res


@app.get("/api/projects/{project_id}/sync-config")
def get_sync_config(project_id: int, user=Depends(_proj_viewer)):
    conn = get_db()
    row = conn.execute("SELECT sync_mode, sync_cron, last_sync_at, last_sync_result,"
                       " git_url FROM projects WHERE id=?", (project_id,)).fetchone()
    counts = conn.execute(
        "SELECT COUNT(*) AS total,"
        " SUM(CASE WHEN source_path IS NOT NULL THEN 1 ELSE 0 END) AS synced,"
        " SUM(CASE WHEN sync_status='missing' THEN 1 ELSE 0 END)   AS missing"
        " FROM test_cases WHERE project_id=?", (project_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404)
    import json as _json
    try:
        last = _json.loads(row["last_sync_result"]) if row["last_sync_result"] else None
    except ValueError:
        last = None
    return {"sync_mode": row["sync_mode"] or "manual",
            "sync_cron": row["sync_cron"],
            "last_sync_at": row["last_sync_at"],
            "last_sync": last,
            "has_git": bool(row["git_url"]),
            "counts": {"total": counts["total"] or 0,
                       "synced": counts["synced"] or 0,
                       "missing": counts["missing"] or 0}}


@app.put("/api/projects/{project_id}/sync-config")
def put_sync_config(project_id: int, req: SyncConfigReq, user=Depends(_proj_admin)):
    if req.sync_mode not in ("manual", "git"):
        raise HTTPException(400, "sync_mode must be 'manual' or 'git'")
    if req.sync_cron:
        try:
            build_trigger(req.sync_cron)
        except Exception as exc:                       # noqa: BLE001
            raise HTTPException(400, f"Invalid sync schedule: {exc}")
    conn = get_db()
    conn.execute("UPDATE projects SET sync_mode=?, sync_cron=? WHERE id=?",
                 (req.sync_mode, req.sync_cron or None, project_id))
    conn.commit()
    conn.close()
    _reload_sync_jobs()
    audit(user, "project.sync_config", project_id=project_id,
          sync_mode=req.sync_mode, sync_cron=req.sync_cron)
    return get_sync_config(project_id, user)


@app.post("/api/projects/{project_id}/sync")
async def sync_test_cases(project_id: int, dry_run: bool = False,
                          user=Depends(_proj_tester)):
    """Reconcile the test case list against the .robot files on disk.

    dry_run reports exactly what would change without writing anything — worth
    running first on a project with existing hand-created cases.
    """
    res = await asyncio.to_thread(_run_tc_sync, project_id, user["username"], dry_run)
    res["summary"] = git_sync.summary(res)
    return res


# ── Project Members ──────────────────────────────────────────────
class MemberAdd(BaseModel):
    username:     str
    project_role: str = "viewer"


@app.get("/api/projects/{project_id}/members")
def list_members(project_id: int, user=Depends(_proj_viewer)):
    conn = get_db()
    rows = conn.execute("""
        SELECT u.id, u.username, u.full_name, u.email, pm.project_role
        FROM project_members pm JOIN users u ON pm.user_id = u.id
        WHERE pm.project_id=? ORDER BY u.username
    """, (project_id,)).fetchall()
    conn.close()
    return rows_to_list(rows)


class MemberBulkAdd(BaseModel):
    usernames:    list[str]
    project_role: str = "viewer"


@app.post("/api/projects/{project_id}/members/bulk")
def add_members_bulk(project_id: int, req: MemberBulkAdd, user=Depends(_proj_admin)):
    """Add several users to a project at one role, in a single transaction."""
    if req.project_role not in ("viewer", "tester", "project_admin"):
        raise HTTPException(400, "project_role must be viewer | tester | project_admin")
    names = [n.strip() for n in dict.fromkeys(req.usernames) if n.strip()]
    if not names:
        raise HTTPException(400, "No usernames supplied")

    conn = get_db()
    marks = ",".join("?" * len(names))
    found = {r["username"]: r["id"] for r in conn.execute(
        f"SELECT id, username FROM users WHERE username IN ({marks})", names).fetchall()}
    missing = [n for n in names if n not in found]
    if missing:
        conn.close()
        raise HTTPException(404, f"Unknown user(s): {', '.join(missing[:10])}")

    for uid in found.values():
        conn.execute(
            "INSERT OR REPLACE INTO project_members (project_id, user_id, project_role) VALUES (?,?,?)",
            (project_id, uid, req.project_role))
    conn.commit()
    conn.close()
    audit(user, "member.add_bulk", project_id=project_id,
          usernames=list(found), role=req.project_role)
    return {"added": len(found), "role": req.project_role}


@app.post("/api/projects/{project_id}/members")
def add_member(project_id: int, req: MemberAdd, user=Depends(_proj_admin)):
    if req.project_role not in ("viewer", "tester", "project_admin"):
        raise HTTPException(400, "project_role must be viewer | tester | project_admin")
    conn = get_db()
    u = conn.execute("SELECT id FROM users WHERE username=?", (req.username,)).fetchone()
    if not u:
        conn.close()
        raise HTTPException(404, f"User '{req.username}' not found")
    try:
        conn.execute(
            "INSERT OR REPLACE INTO project_members (project_id, user_id, project_role) VALUES (?,?,?)",
            (project_id, u["id"], req.project_role),
        )
        conn.commit()
    finally:
        conn.close()
    audit(user, "member.add", project_id=project_id, target=req.username,
          role=req.project_role)
    return {"ok": True}


@app.put("/api/projects/{project_id}/members/{uid}")
def update_member(project_id: int, uid: int, req: MemberAdd, user=Depends(_proj_admin)):
    if req.project_role not in ("viewer", "tester", "project_admin"):
        raise HTTPException(400, "Invalid project_role")
    conn = get_db()
    prev = conn.execute("SELECT project_role FROM project_members"
                        " WHERE project_id=? AND user_id=?", (project_id, uid)).fetchone()
    who = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    conn.execute("UPDATE project_members SET project_role=? WHERE project_id=? AND user_id=?",
                 (req.project_role, project_id, uid))
    conn.commit()
    conn.close()
    audit(user, "member.role_change", project_id=project_id,
          target=who["username"] if who else str(uid),
          **{"from": prev["project_role"] if prev else None, "to": req.project_role})
    return {"ok": True}


@app.delete("/api/projects/{project_id}/members/{uid}")
def remove_member(project_id: int, uid: int, user=Depends(_proj_admin)):
    conn = get_db()
    who = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    conn.execute("DELETE FROM project_members WHERE project_id=? AND user_id=?", (project_id, uid))
    conn.commit()
    conn.close()
    audit(user, "member.remove", project_id=project_id,
          target=who["username"] if who else str(uid))
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════
# SCRIPTS / FILES
# ══════════════════════════════════════════════════════════════════

def _contained(base: Path, target: Path) -> bool:
    """True only if `target` is `base` or sits inside it.

    A plain str.startswith() is NOT sufficient: base '/opt/rf/suites/1' would
    match '/opt/rf/suites/12/...', letting one project reach another's files.
    """
    try:
        target.relative_to(base)
        return True
    except ValueError:
        return False


# Conservative allowlist. Blocks control chars and the HTML/JS metacharacters
# (< > " ' `) that would otherwise ride a filename into the UI.
_SAFE_PATH_RE = re.compile(r"^[A-Za-z0-9._ ()\[\]+&,/-]+$")


def _validate_rel_path(rel: str) -> None:
    if not rel or len(rel) > 400:
        raise HTTPException(400, "Path is empty or too long")
    if not _SAFE_PATH_RE.match(rel):
        raise HTTPException(
            400, "Path may only contain letters, digits, spaces and . _ - ( ) [ ] + & , /")


def _safe_path(project_id: int, rel: str) -> Path:
    _validate_rel_path(rel)
    base   = _project_suites(project_id).resolve()
    target = (base / rel).resolve()
    if not _contained(base, target):
        raise HTTPException(400, "Path traversal denied")
    return target


@app.get("/api/projects/{project_id}/base-path")
def get_base_path(project_id: int, user=Depends(_proj_viewer)):
    """Container-absolute path to this project's suites dir (for locating a file on disk/PVC)."""
    return {"base_path": str(_project_suites(project_id).resolve())}


@app.get("/api/projects/{project_id}/files")
def list_files(project_id: int, user=Depends(_proj_viewer)):
    suites_dir = _project_suites(project_id)
    if not suites_dir.exists():
        return {}
    tree = {}
    base = str(suites_dir)
    for dirpath, dirnames, filenames in os.walk(base):
        dirnames.sort()
        rel_dir = os.path.relpath(dirpath, base).replace("\\", "/")
        folder = "" if rel_dir == "." else rel_dir
        for name in sorted(filenames):
            path = (folder + "/" + name) if folder else name
            tree.setdefault(folder, []).append({"name": name, "path": path})
    return tree


@app.get("/api/projects/{project_id}/files/download-all")
async def download_all_files(project_id: int, user=Depends(_proj_viewer)):
    """Zip the whole project scripts tree and stream it to the browser."""
    suites_dir = _project_suites(project_id)
    if not suites_dir.exists():
        raise HTTPException(404, "No scripts to export")

    project = get_db().execute(
        "SELECT name FROM projects WHERE id=?", (project_id,)).fetchone()
    proj_name = re.sub(r"[^A-Za-z0-9._-]+", "_", (project["name"] if project else str(project_id)))
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"brace-{proj_name}-scripts-{stamp}.zip"

    def build_zip() -> io.BytesIO:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for path in suites_dir.rglob("*"):
                if path.is_file():
                    zf.write(path, path.relative_to(suites_dir))
        buf.seek(0)
        return buf

    # Zipping thousands of files is CPU/IO-bound — build off the event loop
    # so one export doesn't stall every other request on the pod.
    buf = await asyncio.to_thread(build_zip)

    def stream():
        while chunk := buf.read(65536):
            yield chunk

    return StreamingResponse(
        stream(), media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/projects/{project_id}/excel/{filepath:path}")
def read_excel(project_id: int, filepath: str, user=Depends(_proj_viewer)):
    import csv, io
    import openpyxl
    path = _safe_path(project_id, filepath)
    if not path.exists():
        raise HTTPException(404, "File not found")
    ext = path.suffix.lower()
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets = {}
        ROW_LIMIT = 2000
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([("" if c is None else str(c)) for c in row])
                if len(rows) >= ROW_LIMIT:
                    break
            sheets[name] = rows
        wb.close()
        return {"type": "xlsx", "sheets": sheets, "sheet_names": list(wb.sheetnames)}
    elif ext == ".csv":
        text = path.read_text(errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [list(r) for r in reader]
        return {"type": "csv", "sheets": {"Sheet1": rows}, "sheet_names": ["Sheet1"]}
    raise HTTPException(400, "Not an Excel/CSV file")


@app.put("/api/projects/{project_id}/excel/{filepath:path}")
def write_excel(project_id: int, filepath: str, body: dict, user=Depends(_proj_tester)):
    import csv, io
    import openpyxl
    path = _safe_path(project_id, filepath)
    ext = path.suffix.lower()
    sheets: dict = body.get("sheets", {})
    path.parent.mkdir(parents=True, exist_ok=True)
    if ext == ".xlsx":
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row if row else [""])
        wb.save(path)
    elif ext == ".csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        rows = next(iter(sheets.values()), [])
        for row in rows:
            writer.writerow(row)
        path.write_text(buf.getvalue())
    return {"saved": filepath}


@app.get("/api/projects/{project_id}/files/{filepath:path}")
def read_file(project_id: int, filepath: str, user=Depends(_proj_viewer)):
    path = _safe_path(project_id, filepath)
    if not path.exists() or not path.is_file():
        raise HTTPException(404, "File not found")
    return PlainTextResponse(path.read_text(errors="replace"))


@app.put("/api/projects/{project_id}/files/{filepath:path}")
def save_file(project_id: int, filepath: str, body: dict, user=Depends(_proj_tester)):
    path = _safe_path(project_id, filepath)
    if path.suffix not in ALLOWED_EXTS:
        raise HTTPException(400, f"Extension {path.suffix} not allowed")
    path.parent.mkdir(parents=True, exist_ok=True)
    existed = path.exists()
    path.write_text(body.get("content", ""))
    # Editing a .robot file changes what the tests actually do, so this belongs
    # in the trail even though saves are frequent — retention keeps it bounded.
    audit(user, "script.save", project_id=project_id, target=filepath,
          created=not existed, bytes=len(body.get("content", "")))
    return {"saved": filepath}


@app.post("/api/projects/{project_id}/files/upload")
async def upload_files(project_id: int, files: list[UploadFile] = File(...), user=Depends(_proj_tester)):
    suites_dir = _project_suites(project_id)
    suites_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    for uf in files:
        filename = Path(uf.filename).name
        if not filename:
            continue
        data = await uf.read()
        if filename.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                for member in zf.infolist():
                    if member.is_dir():
                        continue
                    mp = Path(member.filename)
                    if mp.suffix not in ALLOWED_EXTS or ".git" in str(mp):
                        continue
                    # zip-slip guard — must be relative_to, not startswith
                    dest = (suites_dir / mp).resolve()
                    if not _contained(suites_dir.resolve(), dest):
                        continue
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    dest.write_bytes(zf.read(member))
                    saved.append(str(mp))
        elif Path(filename).suffix in ALLOWED_EXTS:
            dest = suites_dir / filename
            dest.write_bytes(data)
            saved.append(filename)
    audit(user, "script.upload", project_id=project_id, files=len(saved))
    return {"uploaded": saved}


@app.delete("/api/projects/{project_id}/files/{filepath:path}")
def delete_file(project_id: int, filepath: str, user=Depends(_proj_tester)):
    path = _safe_path(project_id, filepath)
    if not path.exists():
        raise HTTPException(404, "File not found")
    path.unlink()
    audit(user, "script.delete", project_id=project_id, target=filepath)
    return {"deleted": filepath}


class RenameReq(BaseModel):
    old_path: str
    new_path: str


@app.post("/api/projects/{project_id}/fs/rename")
def fs_rename(project_id: int, req: RenameReq, user=Depends(_proj_tester)):
    """Rename or move a file/folder within the project suites dir."""
    src = _safe_path(project_id, req.old_path)
    dst = _safe_path(project_id, req.new_path)
    if not src.exists():
        raise HTTPException(404, "Source not found")
    if dst.exists():
        raise HTTPException(409, f"'{req.new_path}' already exists")
    if src.is_file() and dst.suffix not in ALLOWED_EXTS:
        raise HTTPException(400, f"Extension {dst.suffix} not allowed")
    dst.parent.mkdir(parents=True, exist_ok=True)
    src.rename(dst)
    audit(user, "script.rename", project_id=project_id, target=req.new_path,
          **{"from": req.old_path})
    return {"renamed": req.new_path}


class MkdirReq(BaseModel):
    path: str


@app.post("/api/projects/{project_id}/fs/mkdir")
def fs_mkdir(project_id: int, req: MkdirReq, user=Depends(_proj_tester)):
    path = _safe_path(project_id, req.path)
    if path.exists():
        raise HTTPException(409, "Already exists")
    path.mkdir(parents=True)
    # Keep the folder visible in git and in the file listing until it has content
    (path / ".gitkeep").write_text("")
    return {"created": req.path}


@app.delete("/api/projects/{project_id}/fs/rmdir/{dirpath:path}")
def fs_rmdir(project_id: int, dirpath: str, user=Depends(_proj_tester)):
    """Recursively delete a folder. Refuses the project root."""
    path = _safe_path(project_id, dirpath)
    base = _project_suites(project_id).resolve()
    if path == base:
        raise HTTPException(400, "Cannot delete the project root")
    if not path.exists() or not path.is_dir():
        raise HTTPException(404, "Folder not found")
    n = sum(1 for _ in path.rglob("*") if _.is_file())
    shutil.rmtree(path)
    audit(user, "script.rmdir", project_id=project_id, target=dirpath, files_removed=n)
    return {"deleted": dirpath, "files_removed": n}


@app.get("/api/projects/{project_id}/suites")
def list_suites(project_id: int, user=Depends(_proj_viewer)):
    """Return list of .robot file paths relative to project suites dir."""
    suites_dir = _project_suites(project_id)
    if not suites_dir.exists():
        return []
    results = []
    for p in sorted(suites_dir.rglob("*.robot")):
        rel = str(p.relative_to(suites_dir)).replace("\\", "/")
        parts = rel.lower().split("/")
        # only expose test-case robot files — skip resource/library files
        # a file qualifies if any parent folder is named "testcases" (any case)
        if any(part == "testcases" for part in parts[:-1]):
            results.append(rel)
    return results


# ── Quick run (editor "Run File" button) ──────────────────────────

class QuickRunReq(BaseModel):
    suite_path: str
    extra_args: Optional[str] = None

@app.post("/api/projects/{project_id}/quick-run")
async def quick_run(project_id: int, req: QuickRunReq, user=Depends(_proj_tester)):
    suites_dir = _project_suites(project_id)
    target = _safe_path(project_id, req.suite_path)
    if not target.exists():
        raise HTTPException(404, "File not found")

    run_id  = f"qr-{project_id}-{int(datetime.now().timestamp()*1000)}"
    run_dir = _project_results(project_id) / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    audit(user, "run.quick", project_id=project_id, target=req.suite_path)

    cmd = [
        "python", "-m", "robot",
        "--outputdir", str(run_dir),
        "--output",    "output.xml",
        "--log",       "log.html",
        "--report",    "report.html",
        "--pythonpath", str(suites_dir),
        "--variable",  f"BSS_ENV:{BSS_ENV}",
    ]
    if req.extra_args:
        cmd.extend(_split_args(req.extra_args))
    cmd.append(str(target))

    async def stream():
        yield f"[BRACE] Quick run: {req.suite_path}\n[BRACE] CMD: {' '.join(cmd)}\n\n"
        # A quick run is a browser like any other, so it must draw from the same
        # budget — otherwise the cap is advisory and the pod can still be
        # overloaded from the editor. Say so rather than appearing to hang.
        if _tslots().locked():
            yield "[BRACE] All execution slots busy — waiting for one to free up…\n"
        async with _tslots():
            env = {**os.environ, "DISPLAY": ":99"}
            proc = await asyncio.create_subprocess_exec(
                *cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.STDOUT, env=env
            )
            async for line in proc.stdout:
                yield line.decode(errors="replace")
            await proc.wait()
            rc = proc.returncode
        yield f"\n[BRACE] Exit code: {rc} — {'PASS' if rc == 0 else 'FAIL'}\n"
        yield f"[BRACE] Report: /results/{project_id}/{run_id}/report.html\n"

    return StreamingResponse(stream(), media_type="text/plain")


# ══════════════════════════════════════════════════════════════════
# TEST CASES
# ══════════════════════════════════════════════════════════════════

class TCCreate(BaseModel):
    name:        str
    description: Optional[str] = None
    suite_path:  Optional[str] = None
    extra_args:  Optional[str] = None
    tags:        Optional[str] = None


class TCUpdate(BaseModel):
    name:        Optional[str] = None
    description: Optional[str] = None
    suite_path:  Optional[str] = None
    extra_args:  Optional[str] = None
    tags:        Optional[str] = None


@app.get("/api/projects/{project_id}/test-cases")
def list_test_cases(project_id: int, user=Depends(_proj_viewer)):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM test_cases WHERE project_id=? ORDER BY tc_code", (project_id,)
    ).fetchall()
    conn.close()
    return rows_to_list(rows)


@app.get("/api/projects/{project_id}/tags")
def list_tags(project_id: int, user=Depends(_proj_viewer)):
    """Every tag in use in this project, with how many cases carry it.

    Drives the filter dropdown and the "run everything tagged X" selector, so
    testers pick from what exists instead of guessing at spelling.
    """
    conn = get_db()
    rows = conn.execute(
        "SELECT tags FROM test_cases WHERE project_id=? AND COALESCE(tags,'') != ''",
        (project_id,)).fetchall()
    conn.close()
    counts: dict = {}
    for r in rows:
        for t in (r["tags"] or "").strip(",").split(","):
            if t:
                counts[t] = counts.get(t, 0) + 1
    return [{"tag": t, "count": n} for t, n in sorted(counts.items())]


@app.post("/api/projects/{project_id}/test-cases")
def create_test_case(project_id: int, req: TCCreate, user=Depends(_proj_tester)):
    conn = get_db()
    tc_code = next_tc_code(conn)
    conn.execute(
        "INSERT INTO test_cases (tc_code, project_id, name, description, suite_path,"
        " extra_args, tags) VALUES (?,?,?,?,?,?,?)",
        (tc_code, project_id, req.name, req.description, req.suite_path, req.extra_args,
         _norm_tags(req.tags or "")),
    )
    conn.commit()
    uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = conn.execute("SELECT * FROM test_cases WHERE id=?", (uid,)).fetchone()
    conn.close()
    audit(user, "tc.create", project_id=project_id, target=tc_code, name=req.name)
    return row_to_dict(row)


@app.post("/api/projects/{project_id}/test-cases/bulk-csv")
async def bulk_create_test_cases(project_id: int, file: UploadFile = File(...), user=Depends(_proj_tester)):
    """CSV: name, description (opt), suite_path (opt), extra_args (opt),
    suite (opt, pipe-separated), tags (opt, comma/space separated)"""
    import csv
    content = (await file.read()).decode("utf-8-sig")
    reader  = csv.DictReader(io.StringIO(content))
    if reader.fieldnames:
        reader.fieldnames = [ (f or "").strip().lower() for f in reader.fieldnames ]
    created = []
    conn    = get_db()

    # Suite name → group id. Existing suites are reused; new ones created on demand.
    suite_cache: dict = {}
    suites_created: list = []

    def _resolve_suite(suite_name: str) -> Optional[int]:
        key = suite_name.casefold()
        if key in suite_cache:
            return suite_cache[key]
        row = conn.execute(
            "SELECT id FROM test_groups WHERE project_id=? AND name=? COLLATE NOCASE",
            (project_id, suite_name),
        ).fetchone()
        if row:
            gid = row["id"]
        else:
            conn.execute(
                "INSERT INTO test_groups (project_id, name, description) VALUES (?,?,?)",
                (project_id, suite_name, "Created by bulk CSV upload"),
            )
            conn.commit()
            gid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            suites_created.append(suite_name)
        suite_cache[key] = gid
        return gid

    for raw_row in reader:
        row = { (k or "").strip().lower(): (v.strip() if isinstance(v, str) else v) for k, v in raw_row.items() }
        name = (row.get("name") or "").strip()
        if not name:
            continue
        tc_code = next_tc_code(conn)
        conn.execute(
            "INSERT INTO test_cases (tc_code, project_id, name, description, suite_path,"
            " extra_args, tags) VALUES (?,?,?,?,?,?,?)",
            (tc_code, project_id, name,
             (row.get("description") or "").strip() or None,
             (row.get("suite_path")  or "").strip() or None,
             (row.get("extra_args")  or "").strip() or None,
             _norm_tags(row.get("tags") or "")),
        )
        conn.commit()
        uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Optional `suite` column — one TC may belong to several suites, pipe-separated
        assigned = []
        raw_suites = (row.get("suite") or row.get("suites") or "").strip()
        for suite_name in [s.strip() for s in raw_suites.split("|") if s.strip()]:
            gid = _resolve_suite(suite_name)
            order = conn.execute(
                "SELECT COALESCE(MAX(order_idx), -1) + 1 AS n FROM group_test_cases WHERE group_id=?",
                (gid,),
            ).fetchone()["n"]
            conn.execute(
                "INSERT OR REPLACE INTO group_test_cases (group_id, test_case_id, order_idx) VALUES (?,?,?)",
                (gid, uid, order),
            )
            conn.commit()
            assigned.append(suite_name)

        created.append({"id": uid, "tc_code": tc_code, "name": name, "suites": assigned})

    conn.close()
    audit(user, "tc.bulk_create", project_id=project_id, created=len(created),
          suites_created=suites_created, file=file.filename)
    return {
        "created":        len(created),
        "suites_created": suites_created,
        "suites_used":    len(suite_cache),
        "test_cases":     created,
    }


@app.get("/api/test-cases/{tc_id}/history")
def test_case_history(tc_id: int, limit: int = 50, user=Depends(_current_user)):
    """Execution timeline for one test case, newest first, plus summary stats.

    Answers the question the aggregate flakiness list cannot: has this been
    failing forever, or did it start failing on a particular date?
    """
    conn = get_db()
    tc = conn.execute(
        "SELECT id, tc_code, name, project_id FROM test_cases WHERE id=?", (tc_id,)).fetchone()
    if not tc:
        conn.close()
        raise HTTPException(404, "Test case not found")
    if not _get_project_role(tc["project_id"], user["username"], user["role"]):
        conn.close()
        raise HTTPException(403)

    rows = rows_to_list(conn.execute("""
        SELECT tri.status, tri.started_at, tri.finished_at, tri.rf_run_id,
               tr.run_id, tr.run_name, tr.triggered_by, tr.started_at AS run_started
        FROM test_run_items tri
        JOIN test_runs tr ON tr.run_id = tri.run_id
        WHERE tri.test_case_id=? AND tr.project_id=?
        ORDER BY tr.started_at DESC, tr.id DESC LIMIT ?
    """, (tc_id, tc["project_id"], max(1, min(limit, 500)))).fetchall())
    conn.close()

    results_dir = _project_results(tc["project_id"])
    for r in rows:
        d = None
        if r["started_at"] and r["finished_at"]:
            try:
                d = (datetime.fromisoformat(r["finished_at"])
                     - datetime.fromisoformat(r["started_at"])).total_seconds()
            except ValueError:
                d = None
        r["duration_sec"] = round(d, 1) if d is not None else None
        r["has_log"] = bool(r["rf_run_id"]) and (
            results_dir / r["run_id"] / r["rf_run_id"] / "log.html").exists()

    done = [r for r in rows if r["status"] in ("passed", "failed")]
    passed = sum(1 for r in done if r["status"] == "passed")
    durs = [r["duration_sec"] for r in done if r["duration_sec"] is not None]

    # Consecutive same-status streak from the newest end
    streak, streak_status = 0, None
    for r in done:
        if streak_status is None:
            streak_status, streak = r["status"], 1
        elif r["status"] == streak_status:
            streak += 1
        else:
            break

    # Walk oldest→newest to find where the current failing period began
    first_failure = None
    if streak_status == "failed":
        for r in reversed(done[:streak]):
            first_failure = r["run_started"]
            break

    return {
        "test_case": {"id": tc["id"], "tc_code": tc["tc_code"], "name": tc["name"]},
        "history": rows,
        "stats": {
            "executions":     len(done),
            "passed":         passed,
            "failed":         len(done) - passed,
            "pass_rate":      round(passed / len(done) * 100, 1) if done else None,
            "streak":         streak,
            "streak_status":  streak_status,
            "failing_since":  first_failure,
            "avg_duration":   round(sum(durs) / len(durs), 1) if durs else None,
            "last_status":    done[0]["status"] if done else None,
        },
    }


@app.put("/api/test-cases/{tc_id}")
def update_test_case(tc_id: int, req: TCUpdate, user=Depends(_current_user)):
    conn = get_db()
    tc = conn.execute("SELECT project_id, tc_code, source_path FROM test_cases WHERE id=?",
                      (tc_id,)).fetchone()
    if not tc:
        conn.close()
        raise HTTPException(404)
    try:
        _require_cap(tc["project_id"], user, "edit")
    except HTTPException:
        conn.close()
        raise
    for field, val in [("name", req.name), ("description", req.description),
                       ("suite_path", req.suite_path), ("extra_args", req.extra_args),
                       # normalised so filtering stays exact regardless of how
                       # the tester typed them ("@Smoke, WIP" -> ",smoke,wip,")
                       ("tags", None if req.tags is None else _norm_tags(req.tags))]:
        if val is not None:
            conn.execute(f"UPDATE test_cases SET {field}=? WHERE id=?", (val, tc_id))
    conn.commit()
    row = conn.execute("SELECT * FROM test_cases WHERE id=?", (tc_id,)).fetchone()
    conn.close()
    audit(user, "tc.update", project_id=tc["project_id"], target=tc["tc_code"],
          fields=[f for f in ("name", "description", "suite_path", "extra_args", "tags")
                  if getattr(req, f) is not None],
          git_synced=bool(tc["source_path"]))
    return row_to_dict(row)


@app.delete("/api/test-cases/{tc_id}")
def delete_test_case(tc_id: int, user=Depends(_current_user)):
    conn = get_db()
    tc = conn.execute("SELECT project_id, tc_code FROM test_cases WHERE id=?",
                      (tc_id,)).fetchone()
    if not tc:
        conn.close()
        raise HTTPException(404)
    try:
        _require_cap(tc["project_id"], user, "manage")
    except HTTPException:
        conn.close()
        raise
    # Detach run history — test_run_items.test_case_id has no ON DELETE clause.
    # tc_code/tc_name are denormalised on the item, so past runs stay readable.
    conn.execute("UPDATE test_run_items SET test_case_id=NULL WHERE test_case_id=?", (tc_id,))
    conn.execute("DELETE FROM test_cases WHERE id=?", (tc_id,))
    conn.commit()
    conn.close()
    audit(user, "tc.delete", project_id=tc["project_id"], target=tc["tc_code"])
    return {"deleted": tc_id}


# ══════════════════════════════════════════════════════════════════
# TEST GROUPS (SUITES)
# ══════════════════════════════════════════════════════════════════

class GroupCreate(BaseModel):
    name:        str
    description: Optional[str] = None


class GroupTCAdd(BaseModel):
    test_case_id: int
    order_idx:    int = 0


@app.get("/api/projects/{project_id}/groups")
def list_groups(project_id: int, user=Depends(_proj_viewer)):
    conn = get_db()
    groups = conn.execute(
        "SELECT * FROM test_groups WHERE project_id=? ORDER BY name", (project_id,)
    ).fetchall()
    result = []
    for g in groups:
        members = conn.execute("""
            SELECT tc.id, tc.tc_code, tc.name, tc.description, tc.suite_path,
                   tc.last_run_status, gtc.order_idx
            FROM test_cases tc
            JOIN group_test_cases gtc ON tc.id = gtc.test_case_id
            WHERE gtc.group_id=? ORDER BY gtc.order_idx
        """, (g["id"],)).fetchall()
        d = dict(g)
        d["test_cases"] = rows_to_list(members)
        d["tc_count"]   = len(d["test_cases"])
        result.append(d)
    conn.close()
    return result


@app.post("/api/projects/{project_id}/groups")
def create_group(project_id: int, req: GroupCreate, user=Depends(_proj_tester)):
    conn = get_db()
    conn.execute("INSERT INTO test_groups (project_id, name, description) VALUES (?,?,?)",
                 (project_id, req.name, req.description))
    conn.commit()
    uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = conn.execute("SELECT * FROM test_groups WHERE id=?", (uid,)).fetchone()
    conn.close()
    audit(user, "suite.create", project_id=project_id, target=req.name)
    return row_to_dict(row)


@app.put("/api/groups/{gid}")
def update_group(gid: int, req: GroupCreate, user=Depends(_current_user)):
    conn = get_db()
    g = _owned_row(conn, "SELECT project_id, name FROM test_groups WHERE id=?",
                   gid, user, "edit", "Suite not found")
    conn.execute("UPDATE test_groups SET name=?, description=? WHERE id=?",
                 (req.name, req.description, gid))
    conn.commit()
    conn.close()
    audit(user, "suite.update", project_id=g["project_id"], target=req.name,
          renamed_from=g["name"] if g["name"] != req.name else None)
    return {"ok": True}


@app.delete("/api/groups/{gid}")
def delete_group(gid: int, user=Depends(_current_user)):
    conn = get_db()
    # 'edit', not 'manage': testers have always been able to delete a suite, and
    # a suite carries no history of its own — the runs it produced survive.
    g = _owned_row(conn, "SELECT project_id, name FROM test_groups WHERE id=?",
                   gid, user, "edit", "Suite not found")
    # Detach historical runs first — test_runs.group_id has no ON DELETE clause,
    # so the FK would block the delete. Runs keep their own name/stats, so
    # nulling the link preserves history instead of cascading it away.
    conn.execute("UPDATE test_runs SET group_id=NULL WHERE group_id=?", (gid,))
    conn.execute("DELETE FROM schedules WHERE group_id=?", (gid,))
    conn.execute("DELETE FROM test_groups WHERE id=?", (gid,))
    conn.commit()
    conn.close()
    audit(user, "suite.delete", project_id=g["project_id"], target=g["name"])
    return {"ok": True}


class GroupTCBulkAdd(BaseModel):
    test_case_ids: list[int]


@app.post("/api/groups/{gid}/test-cases/bulk")
def add_tcs_to_group(gid: int, req: GroupTCBulkAdd, user=Depends(_current_user)):
    """Add many test cases to a suite in one transaction.

    The per-TC endpoint needed one HTTP round trip each, so adding a 75-case
    regression pack meant 75 requests and a partially-filled suite if one failed.
    """
    conn = get_db()
    g = _owned_row(conn, "SELECT project_id, name FROM test_groups WHERE id=?",
                   gid, user, "edit", "Suite not found")

    ids = list(dict.fromkeys(req.test_case_ids))       # de-dupe, keep order
    if not ids:
        conn.close()
        raise HTTPException(400, "No test cases supplied")

    # Every id must belong to this suite's project
    marks = ",".join("?" * len(ids))
    owned = {r["id"] for r in conn.execute(
        f"SELECT id FROM test_cases WHERE id IN ({marks}) AND project_id=?",
        (*ids, g["project_id"])).fetchall()}
    rejected = [i for i in ids if i not in owned]
    if rejected:
        conn.close()
        raise HTTPException(404, f"{len(rejected)} test case(s) not found in this project")

    start = conn.execute(
        "SELECT COALESCE(MAX(order_idx), -1) + 1 AS n FROM group_test_cases WHERE group_id=?",
        (gid,)).fetchone()["n"]
    already = {r["test_case_id"] for r in conn.execute(
        "SELECT test_case_id FROM group_test_cases WHERE group_id=?", (gid,)).fetchall()}

    added = 0
    for tc_id in ids:
        if tc_id in already:
            continue
        conn.execute(
            "INSERT INTO group_test_cases (group_id, test_case_id, order_idx) VALUES (?,?,?)",
            (gid, tc_id, start + added))
        added += 1
    conn.commit()
    conn.close()
    audit(user, "suite.add_cases", project_id=g["project_id"], target=g["name"],
          added=added, requested=len(ids))
    return {"added": added, "skipped_already_present": len(ids) - added}


@app.post("/api/groups/{gid}/test-cases")
def add_tc_to_group(gid: int, req: GroupTCAdd, user=Depends(_current_user)):
    conn = get_db()
    g = _owned_row(conn, "SELECT project_id, name FROM test_groups WHERE id=?",
                   gid, user, "edit", "Suite not found")
    # The test case must live in the same project as the suite
    owned = conn.execute("SELECT 1 FROM test_cases WHERE id=? AND project_id=?",
                         (req.test_case_id, g["project_id"])).fetchone()
    if not owned:
        conn.close()
        raise HTTPException(404, "Test case not found in this project")
    conn.execute(
        "INSERT OR REPLACE INTO group_test_cases (group_id, test_case_id, order_idx) VALUES (?,?,?)",
        (gid, req.test_case_id, req.order_idx),
    )
    conn.commit()
    conn.close()
    audit(user, "suite.add_cases", project_id=g["project_id"], target=g["name"], added=1)
    return {"ok": True}


@app.delete("/api/groups/{gid}/test-cases/{tc_id}")
def remove_tc_from_group(gid: int, tc_id: int, user=Depends(_current_user)):
    conn = get_db()
    g = _owned_row(conn, "SELECT project_id, name FROM test_groups WHERE id=?",
                   gid, user, "edit", "Suite not found")
    conn.execute("DELETE FROM group_test_cases WHERE group_id=? AND test_case_id=?", (gid, tc_id))
    conn.commit()
    conn.close()
    audit(user, "suite.remove_case", project_id=g["project_id"], target=g["name"],
          test_case_id=tc_id)
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════
# TEST RUNS — Unified Engine
# ══════════════════════════════════════════════════════════════════

class RunRequest(BaseModel):
    tc_ids:    Optional[list[int]] = None   # specific TCs
    group_id:  Optional[int]       = None   # or an entire group
    tag:       Optional[str]       = None   # or every TC carrying this tag
    run_name:  Optional[str]       = None
    extra_args: Optional[str]      = None
    # Cases in flight at once for this run. Omit for the server default; set 1
    # for suites whose cases depend on each other's order.
    parallel:  Optional[int]       = None


@app.post("/api/projects/{project_id}/runs")
async def trigger_run(project_id: int, req: RunRequest, user=Depends(_proj_tester)):
    conn = get_db()

    # Resolve test cases
    if req.group_id:
        # The suite must belong to the project the caller is authorised for,
        # otherwise a tester could run another project's suite via its id.
        group_name = conn.execute(
            "SELECT name FROM test_groups WHERE id=? AND project_id=?",
            (req.group_id, project_id)).fetchone()
        if not group_name:
            conn.close()
            raise HTTPException(404, "Suite not found in this project")
        members = conn.execute("""
            SELECT tc.* FROM test_cases tc
            JOIN group_test_cases gtc ON tc.id = gtc.test_case_id
            WHERE gtc.group_id=? AND tc.project_id=? ORDER BY gtc.order_idx
        """, (req.group_id, project_id)).fetchall()
        tcs = rows_to_list(members)
        default_name = f"{group_name['name']} run"
    elif req.tc_ids:
        placeholders = ",".join("?" * len(req.tc_ids))
        members = conn.execute(
            f"SELECT * FROM test_cases WHERE id IN ({placeholders}) AND project_id=?",
            (*req.tc_ids, project_id),
        ).fetchall()
        tcs = rows_to_list(members)
        default_name = f"Ad-hoc run ({len(tcs)} TCs)"
    elif req.tag:
        # Tags cut across suites — this is the selection a suite cannot express.
        tag = _norm_tag(req.tag)
        if not tag:
            conn.close()
            raise HTTPException(400, "Invalid tag")
        # tags are stored already wrapped in commas (",smoke,wip,"), so this
        # matches a whole tag — "smoke" will not match "smoketest".
        members = conn.execute(
            "SELECT * FROM test_cases WHERE project_id=? AND COALESCE(tags,'') LIKE ?"
            " ORDER BY tc_code", (project_id, f"%,{tag},%")).fetchall()
        tcs = rows_to_list(members)
        default_name = f"Tag run: {tag}"
    else:
        conn.close()
        raise HTTPException(400, "Provide tc_ids, group_id or tag")

    if not tcs:
        conn.close()
        raise HTTPException(400, "No test cases found")

    conn.close()
    out = _start_run(project_id, tcs, req.run_name or default_name,
                     user["username"], req.extra_args, group_id=req.group_id,
                     parallel=req.parallel)
    audit(user, "run.trigger", project_id=project_id, target=out["run_id"],
          total=out["total"], group_id=req.group_id, tag=req.tag,
          parallel=out["parallel"])
    return out


def _start_run(project_id: int, tcs: list, run_name: str, username: str,
               extra_args: Optional[str], group_id: Optional[int] = None,
               rerun_of: Optional[str] = None, parallel: Optional[int] = None) -> dict:
    """Persist a run + its items and hand it to the executor.

    Shared by the normal trigger, re-run-failed and the scheduler so the three
    paths cannot drift apart.
    """
    conn = get_db()
    run_id = f"run-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"

    # Write started_at explicitly: the column default is SQLite datetime('now')
    # which is UTC, while finished_at is written with Python datetime.now()
    # (local). Mixing the two inflated every computed duration by the UTC offset.
    conn.execute(
        "INSERT INTO test_runs (run_id, project_id, group_id, run_name, triggered_by, total,"
        " started_at, status, rerun_of) VALUES (?,?,?,?,?,?,?,?,?)",
        (run_id, project_id, group_id, run_name, username, len(tcs),
         datetime.now().isoformat(timespec="seconds"), "queued", rerun_of),
    )
    conn.commit()

    items = []
    for tc in tcs:
        conn.execute(
            "INSERT INTO test_run_items (run_id, test_case_id, tc_code, tc_name, status) VALUES (?,?,?,?,?)",
            (run_id, tc["id"], tc["tc_code"], tc["name"], "pending"),
        )
        conn.commit()
        item_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        items.append({"item_id": item_id, "tc": tc})
    conn.close()

    # Starts queued; _execute_run flips it to running once a slot frees up.
    _active_runs[run_id] = {"status": "queued", "total": len(tcs), "passed": 0, "failed": 0}
    _metrics["runs_started"] += 1
    log.info("Run queued", extra={"run_id": run_id, "project_id": project_id,
                                  "user": username, "total": len(tcs),
                                  "rerun_of": rerun_of})
    width = max(1, min(MAX_CONCURRENT_TESTS, parallel or RUN_PARALLEL_DEFAULT, len(tcs)))
    asyncio.create_task(_execute_run(run_id, project_id, items, extra_args, width))

    queued_ahead = max(0, sum(1 for r in _active_runs.values() if r["status"] == "queued") - 1)
    return {"run_id": run_id, "run_name": run_name, "total": len(tcs),
            "status": "queued", "queued_ahead": queued_ahead, "rerun_of": rerun_of,
            "parallel": width}


class RerunReq(BaseModel):
    include_cancelled: bool = False   # also retry items stopped by cancel/timeout


@app.post("/api/runs/{run_id}/rerun-failed")
async def rerun_failed(run_id: str, req: RerunReq, user=Depends(_current_user)):
    """Re-run only the test cases that failed in a previous run.

    Each BRACE test case is its own robot invocation, so this is simply a new
    run over the failed subset — no --rerunfailed bookkeeping required.
    """
    conn = get_db()
    tr = conn.execute(
        "SELECT project_id, run_name FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    if not tr:
        conn.close()
        raise HTTPException(404, "Run not found")
    project_id = tr["project_id"]
    try:
        _require_cap(project_id, user, "run")
    except HTTPException:
        conn.close()
        raise

    wanted = ["failed"] + (["cancelled"] if req.include_cancelled else [])
    marks = ",".join("?" * len(wanted))
    # Join to test_cases so a case deleted since the original run is skipped
    # rather than producing a run with nothing to execute.
    tcs = rows_to_list(conn.execute(
        f"""SELECT DISTINCT tc.* FROM test_run_items tri
            JOIN test_cases tc ON tc.id = tri.test_case_id
            WHERE tri.run_id=? AND tri.status IN ({marks}) AND tc.project_id=?
            ORDER BY tc.tc_code""",
        (run_id, *wanted, project_id)).fetchall())
    conn.close()

    if not tcs:
        raise HTTPException(400, "Nothing to re-run — no failed test cases in that run")

    base = tr["run_name"] or run_id
    base = re.sub(r"\s*\(retry \d+\)$", "", base)      # don't stack "(retry 1) (retry 1)"
    out = _start_run(project_id, tcs, f"{base} (retry {len(tcs)})",
                     user["username"], None, rerun_of=run_id)
    audit(user, "run.rerun_failed", project_id=project_id, target=out["run_id"],
          rerun_of=run_id, total=out["total"])
    return out


async def _execute_run(run_id: str, project_id: int, items: list, extra_args: Optional[str],
                       parallel: Optional[int] = None):
    """Wait for a free execution slot, then run the suite."""
    async with _slots():
        # The run may have been cancelled while it sat in the queue
        if run_id in _cancelled_runs:
            _cancelled_runs.discard(run_id)
            _active_runs.pop(run_id, None)
            log.info("Run %s was cancelled while queued — not starting.", run_id)
            return
        conn = get_db()
        conn.execute("UPDATE test_runs SET status='running', started_at=? WHERE run_id=?",
                     (datetime.now().isoformat(timespec="seconds"), run_id))
        conn.commit()
        conn.close()
        if run_id in _active_runs:
            _active_runs[run_id]["status"] = "running"
        _publish(run_id, "summary", {"status": "running"})
        try:
            await _run_suite(run_id, project_id, items, extra_args, parallel)
        finally:
            _cancelled_runs.discard(run_id)


# ── Failure extraction ───────────────────────────────────────────
# Robot writes everything needed to explain a failure into output.xml. Pulling
# it out here means the run list can say *why* a case failed, instead of making
# someone open log.html and hunt for the red keyword.
_IMG_SRC_RE = re.compile(r'src="([^"]+)"')


def _extract_failure(out_xml: Path, item_dir: Path):
    """Return (summary, detail, screenshot_name) for a failed output.xml.

    summary  — the failing keyword and its library, e.g.
               "Click Element (SeleniumLibrary)", suitable for one table cell.
    detail   — the failure message robot reported.
    screenshot_name — a file in item_dir, or None. Only names an existing file,
               so the UI never renders a broken thumbnail.
    """
    root = ET.parse(str(out_xml)).getroot()

    # Test-level message is the reliable fallback: always present on failure.
    detail = None
    for test in root.iter("test"):
        st = test.find("status")
        if st is not None and st.get("status") == "FAIL":
            detail = (st.text or "").strip() or None
            break

    # The failing keyword is the *innermost* one that failed — outer keywords
    # inherit FAIL from their children, so the outermost would just say
    # "Run Test", which explains nothing.
    kw_name = kw_owner = None
    for kw in root.iter("kw"):
        st = kw.find("status")
        if st is None or st.get("status") != "FAIL":
            continue
        if any(c.find("status") is not None and c.find("status").get("status") == "FAIL"
               for c in kw.findall("kw")):
            continue                                  # a child failed; not the leaf
        kw_name  = kw.get("name") or kw_name
        # 'owner' (RF 7) / 'library' (RF <=6) — worth showing: it separates a
        # SeleniumLibrary timeout from a failure in the team's own resource file.
        kw_owner = kw.get("owner") or kw.get("library") or None
        if not detail:
            for msg in kw.findall("msg"):
                if msg.get("level") in ("FAIL", "ERROR") and (msg.text or "").strip():
                    detail = msg.text.strip()
                    break

    # Robot records no line number on keywords — only on the test — so naming a
    # line here would point at the test header, not the failing step. Show the
    # keyword and its library instead of a number that would mislead.
    summary = kw_name or "Test failed"
    if kw_name and kw_owner:
        summary = f"{kw_name} ({kw_owner})"

    # Selenium logs screenshots as an html msg holding an <img src="...">.
    shot = None
    for msg in root.iter("msg"):
        if msg.get("html") != "true" or not msg.text:
            continue
        m = _IMG_SRC_RE.search(msg.text)
        if not m:
            continue
        name = m.group(1).split("/")[-1]
        # Robot writes screenshots beside output.xml; confirm before pointing
        # the UI at it, and keep it inside item_dir so a crafted src cannot
        # escape into another run's directory.
        cand = (item_dir / name).resolve()
        if _contained(item_dir.resolve(), cand) and cand.is_file():
            shot = name                               # last one = closest to the failure

    if detail and len(detail) > 2000:
        detail = detail[:2000] + " …"
    return summary[:300], detail, shot


# ── Tags ─────────────────────────────────────────────────────────
_TAG_RE = re.compile(r"^[a-z0-9][a-z0-9._-]{0,39}$")


def _norm_tag(tag: str) -> str:
    """Lower-case, strip a leading '@'. Returns '' if it isn't a usable tag."""
    t = (tag or "").strip().lstrip("@").lower()
    return t if _TAG_RE.match(t) else ""


def _norm_tags(raw: str) -> str:
    """Normalise a comma/space separated tag string for storage.

    Stored as ',a,b,c,' — the wrapping commas let a LIKE '%,tag,%' match a whole
    tag without also matching 'smoke' inside 'smoketest'.
    """
    parts = [_norm_tag(p) for p in re.split(r"[,\s]+", raw or "") if p.strip()]
    seen, out = set(), []
    for p in parts:
        if p and p not in seen:
            seen.add(p)
            out.append(p)
    return f",{','.join(out)}," if out else ""


def _split_args(raw: str) -> list:
    """Split robot arguments, honouring quotes.

    Falls back to a plain split when the string has unbalanced quotes, so a
    malformed extra_args degrades the way it always did instead of raising and
    failing the test case outright.
    """
    import shlex
    try:
        return shlex.split(raw)
    except ValueError:
        log.warning("Could not parse extra_args %r — falling back to plain split", raw)
        return raw.split()


def _db_write(statements: list) -> None:
    """Run (sql, params) pairs in one transaction. Blocking — call via to_thread.

    SQLite waits up to busy_timeout for a lock. Executed directly inside an
    async function that wait is spent on the event loop, so with several tests
    finishing at once the whole server stops answering — including /health,
    which used to get the pod restarted mid-run.
    """
    conn = get_db()
    try:
        for sql, params in statements:
            conn.execute(sql, params)
        conn.commit()
    finally:
        conn.close()


async def _run_one_item(run_id: str, project_id: int, item: dict, extra_args: Optional[str],
                        run_dir: Path, suites_dir: Path, tally: dict) -> Optional[str]:
    """Execute one test case. Returns its output.xml path, or None.

    Split out of _run_suite so several can be in flight at once. Every robot
    process — whichever run it belongs to — must hold a slot from the global
    _tslots() budget, because that budget counts browsers, and browsers are what
    exhausts the pod.
    """
    tc        = item["tc"]
    item_id   = item["item_id"]
    rf_run_id = f"{run_id}-tc-{tc['tc_code'] or tc['id']}"
    item_dir  = run_dir / rf_run_id

    async with _tslots():
        # Cancelled (or the whole run torn down) while queued for a slot —
        # don't start a browser for work nobody is waiting for.
        if run_id in _cancelled_runs or run_id not in _active_runs:
            return None

        item_dir.mkdir(parents=True, exist_ok=True)
        await asyncio.to_thread(_db_write, [(
            "UPDATE test_run_items SET rf_run_id=?, status='running', started_at=? WHERE id=?",
            (rf_run_id, datetime.now().isoformat(), item_id))])
        _publish(run_id, "item", {"id": item_id, "tc_code": tc.get("tc_code"),
                                  "tc_name": tc.get("name"), "status": "running",
                                  "rf_run_id": rf_run_id})

        suite_path = tc.get("suite_path")
        target = suites_dir / suite_path if suite_path else suites_dir

        cmd = [
            "python", "-m", "robot",
            "--outputdir",  str(item_dir),
            "--output",     "output.xml",
            "--log",        "log.html",
            "--report",     "report.html",
            "--pythonpath", str(suites_dir),
            "--variable",   f"RUN_ID:{run_id}",
            "--variable",   f"BSS_ENV:{BSS_ENV}",
        ]
        # shlex, not str.split: a git-synced case carries --test "Verify Login",
        # and splitting that on whitespace hands robot two broken fragments.
        if tc.get("extra_args"):
            cmd.extend(_split_args(tc["extra_args"]))
        if extra_args:
            cmd.extend(_split_args(extra_args))
        cmd.append(str(target))

        log_file = item_dir / "console.log"
        env = {**os.environ, "DISPLAY": ":99"}

        timed_out = False
        t_start = time.monotonic()
        with open(log_file, "w") as f:
            proc = await asyncio.create_subprocess_exec(
                *cmd, stdout=f, stderr=subprocess.STDOUT, env=env,
            )
            _active_procs[rf_run_id] = proc
            try:
                # Bounded: a wedged browser would otherwise hold an execution
                # slot indefinitely and stall every queued run behind it.
                exit_code = await asyncio.wait_for(proc.wait(), timeout=TEST_TIMEOUT_SEC)
            except asyncio.TimeoutError:
                timed_out = True
                log.warning("Test %s exceeded %ds — terminating.", rf_run_id, TEST_TIMEOUT_SEC)
                proc.terminate()
                try:
                    exit_code = await asyncio.wait_for(proc.wait(), timeout=15)
                except asyncio.TimeoutError:
                    proc.kill()                      # SIGTERM ignored — force it
                    exit_code = await proc.wait()
                f.write(f"\n\n[BRACE] Timed out after {TEST_TIMEOUT_SEC}s and was terminated.\n")
            finally:
                _active_procs.pop(rf_run_id, None)

    status = "passed" if (exit_code == 0 and not timed_out) else "failed"
    if status == "passed":
        tally["passed"] += 1
        _metrics["tests_passed"] += 1
    else:
        tally["failed"] += 1
        _metrics["tests_failed"] += 1
    if timed_out:
        _metrics["tests_timeout"] += 1
    elapsed = max(0.0, time.monotonic() - t_start)
    _metrics["test_seconds"] += elapsed
    _metrics["test_count"]   += 1
    log.info("Test case finished", extra={
        "run_id": run_id, "project_id": project_id, "tc_code": tc.get("tc_code"),
        "status": status, "duration_sec": round(elapsed, 1),
        "timed_out": timed_out})

    out_xml = item_dir / "output.xml"

    # Pull the failing keyword and screenshot straight out of output.xml so the
    # UI can show why it failed without anyone opening log.html.
    fail_summary = fail_detail = fail_shot = None
    if status == "failed" and out_xml.exists():
        try:
            fail_summary, fail_detail, fail_shot = await asyncio.to_thread(
                _extract_failure, out_xml, item_dir)
        except Exception as exc:                       # noqa: BLE001 — cosmetic
            log.debug("Could not extract failure detail for %s: %s", rf_run_id, exc)

    now = datetime.now().isoformat()
    await asyncio.to_thread(_db_write, [
        ("UPDATE test_run_items SET status=?, finished_at=?, fail_summary=?,"
         " fail_detail=?, fail_screenshot=? WHERE id=?",
         (status, now, fail_summary, fail_detail, fail_shot, item_id)),
        ("UPDATE test_cases SET last_run_status=?, last_run_at=? WHERE id=?",
         (status, now, tc["id"])),
    ])

    # cancel_run removes the entry, so don't assume it is still there
    if run_id in _active_runs:
        _active_runs[run_id]["passed"] = tally["passed"]
        _active_runs[run_id]["failed"] = tally["failed"]

    # One event carries both the case's verdict and the new totals, so a watching
    # browser needs no follow-up request to redraw the header.
    _publish(run_id, "item", {
        "id": item_id, "tc_code": tc.get("tc_code"), "tc_name": tc.get("name"),
        "status": status, "rf_run_id": rf_run_id, "fail_summary": fail_summary,
        "passed": tally["passed"], "failed": tally["failed"]})

    return str(out_xml) if out_xml.exists() else None


async def _run_suite(run_id: str, project_id: int, items: list, extra_args: Optional[str],
                     parallel: Optional[int] = None):
    """Execute the run's test cases, then merge their reports with rebot.

    Cases run `parallel`-wide within the run; the global _tslots() budget still
    bounds how many actually execute at once across all runs.
    """
    run_dir     = _project_results(project_id) / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    suites_dir  = _project_suites(project_id)
    tally       = {"passed": 0, "failed": 0}

    width = max(1, min(MAX_CONCURRENT_TESTS, parallel or RUN_PARALLEL_DEFAULT, len(items)))
    gate  = asyncio.Semaphore(width)

    async def worker(item):
        async with gate:
            # Re-check per item: cancelling should stop cases that have not
            # started yet, not just kill the ones already running.
            if run_id in _cancelled_runs or run_id not in _active_runs:
                return None
            return await _run_one_item(run_id, project_id, item, extra_args,
                                       run_dir, suites_dir, tally)

    log.info("Run executing", extra={"run_id": run_id, "project_id": project_id,
                                     "total": len(items), "parallel": width})
    # Results come back in submission order regardless of completion order, so
    # the merged report keeps the suite's original sequence.
    results = await asyncio.gather(*(worker(i) for i in items), return_exceptions=True)

    output_files = []
    for item, res in zip(items, results):
        if isinstance(res, Exception):
            log.error("Test case raised", extra={"run_id": run_id,
                                                 "tc_code": item["tc"].get("tc_code"),
                                                 "error": str(res)})
            # An exception means no verdict was written; leaving the item
            # 'running' would strand it forever.
            tally["failed"] += 1
            conn = get_db()
            conn.execute("UPDATE test_run_items SET status='failed', finished_at=?,"
                         " fail_summary=? WHERE id=? AND status='running'",
                         (datetime.now().isoformat(), f"Executor error: {res}", item["item_id"]))
            conn.commit()
            conn.close()
        elif res:
            output_files.append(res)

    passed, failed = tally["passed"], tally["failed"]

    if run_id in _cancelled_runs or run_id not in _active_runs:
        log.info("Run %s cancelled — %d/%d case(s) completed.", run_id,
                 passed + failed, len(items))

    # Merge reports with rebot
    final_status = "passed" if failed == 0 else "failed"
    if output_files:
        try:
            rebot_dir = run_dir / "combined"
            rebot_dir.mkdir(exist_ok=True)
            # to_thread: merging many output.xml files is CPU-bound and this
            # coroutine shares the event loop with every HTTP request.
            await asyncio.to_thread(
                subprocess.run,
                ["python", "-m", "robot.rebot",
                 "--outputdir", str(rebot_dir),
                 "--output", "output.xml",
                 "--log",    "log.html",
                 "--report", "report.html",
                 "--name",   run_id,
                 *output_files],
                capture_output=True,
            )
        except Exception as e:
            log.warning("rebot merge failed: %s", e)

    conn = get_db()
    # Don't overwrite a 'cancelled' verdict the user already set
    conn.execute(
        "UPDATE test_runs SET status=?, passed=?, failed=?, finished_at=?"
        " WHERE run_id=? AND status != 'cancelled'",
        (final_status, passed, failed, datetime.now().isoformat(), run_id),
    )
    conn.commit()
    conn.close()
    if run_id in _active_runs:
        _active_runs[run_id]["status"] = final_status
    _metrics["runs_passed" if final_status == "passed" else "runs_failed"] += 1
    log.info("Run finished", extra={"run_id": run_id, "project_id": project_id,
                                    "status": final_status,
                                    "passed": passed, "failed": failed})
    # 'done' tells each watching browser to close its EventSource — without it
    # they would hold an idle connection open until the router timed it out.
    _publish(run_id, "done", {"status": final_status, "passed": passed,
                              "failed": failed})

    # Fire-and-forget: one email for the whole run, never one per failed case.
    # _notify_run_finished swallows everything, so a broken relay cannot turn a
    # completed run into a failed one — and the run is already committed above.
    if run_id not in _cancelled_runs:
        asyncio.create_task(_notify_run_finished(run_id, final_status))


# ── Notification dispatch ────────────────────────────────────────
def _notify_recipients(cfg: dict, run: dict) -> list:
    """Configured list, optionally plus whoever triggered the run."""
    to = mailer.parse_recipients(cfg.get("recipients") or "")
    if cfg.get("notify_triggerer") and run.get("triggered_by"):
        conn = get_db()
        row = conn.execute("SELECT email FROM users WHERE username=?",
                           (run["triggered_by"],)).fetchone()
        conn.close()
        addr = (row["email"] or "").strip() if row else ""
        if addr and addr not in to and mailer.parse_recipients(addr):
            to.append(addr)
    return to


def _should_notify(cfg: dict, run: dict, status: str, scheduled: bool) -> bool:
    """Event filter plus the only_on_change gate.

    only_on_change is what keeps this feature alive: a nightly suite that has
    been failing for three weeks otherwise sends 21 identical emails, everyone
    filters BRACE to a folder, and the next real regression is missed.
    """
    if not cfg.get("enabled"):
        return False

    # Record the outcome FIRST, and unconditionally.
    #
    # Recording used to happen only when we were about to email, which meant a
    # run that passed while on_run_passed was off never updated the state. The
    # stored status stayed 'failed', so the next real failure looked unchanged
    # and was suppressed: fail -> pass -> fail sent exactly one email, and the
    # regression in the third run was never reported.
    scope = f"p{run['project_id']}:" + (f"g{run['group_id']}" if run.get("group_id") else "adhoc")
    conn = get_db()
    row = conn.execute("SELECT last_status FROM notify_state WHERE scope=?", (scope,)).fetchone()
    prev = row["last_status"] if row else None
    conn.execute("INSERT INTO notify_state (scope, last_status, last_sent_at) VALUES (?,?,?)"
                 " ON CONFLICT(scope) DO UPDATE SET last_status=excluded.last_status,"
                 " last_sent_at=excluded.last_sent_at",
                 (scope, status, datetime.now().isoformat(timespec="seconds")))
    conn.commit()
    conn.close()

    wanted = ((status == "failed" and cfg.get("on_scheduled_failed") and scheduled)
              or (status == "failed" and cfg.get("on_run_failed") and not scheduled)
              or (status == "passed" and cfg.get("on_run_passed")))
    if not wanted:
        return False
    if not cfg.get("only_on_change"):
        return True
    if prev == status:
        log.info("Notification suppressed — unchanged outcome",
                 extra={"run_id": run.get("run_id"), "scope": scope, "status": status})
        return False
    return True


async def _notify_run_finished(run_id: str, status: str) -> None:
    """Email one summary for a finished run. Never raises.

    Called from the executor: a broken relay, a DNS failure or a typo in the
    recipient list must not affect the run that just completed.
    """
    try:
        smtp = _smtp_row(decrypt=True)
        if not (smtp.get("enabled") and (smtp.get("host") or "").strip()):
            return
        conn = get_db()
        run = conn.execute(
            "SELECT tr.*, p.name AS project_name FROM test_runs tr"
            " JOIN projects p ON p.id = tr.project_id WHERE tr.run_id=?",
            (run_id,)).fetchone()
        if not run:
            conn.close()
            return
        run = dict(run)
        items = rows_to_list(conn.execute(
            "SELECT tc_code, tc_name, status, fail_summary, fail_detail"
            " FROM test_run_items WHERE run_id=? ORDER BY id", (run_id,)).fetchall())
        conn.close()

        cfg = _notify_cfg(run["project_id"])
        scheduled = (run.get("triggered_by") == "scheduler")
        if not _should_notify(cfg, run, status, scheduled):
            return
        to = _notify_recipients(cfg, run)
        if not to:
            log.warning("Notifications on but no valid recipients",
                        extra={"project_id": run["project_id"], "run_id": run_id})
            return

        if run.get("started_at") and run.get("finished_at"):
            try:
                secs = (datetime.fromisoformat(run["finished_at"])
                        - datetime.fromisoformat(run["started_at"])).total_seconds()
                run["duration_txt"] = (f"{int(secs)}s" if secs < 60
                                       else f"{int(secs // 60)}m {int(secs % 60)}s")
            except ValueError:
                pass

        subject, text, html_body = mailer.run_email(run, items, run["project_name"])
        # One retry: relays drop connections, and losing the only alert about a
        # failed nightly run to a transient blip is exactly what must not happen.
        for attempt in (1, 2):
            try:
                await asyncio.to_thread(mailer.send_mail, smtp, to, subject, text, html_body)
                log.info("Notification sent", extra={"run_id": run_id, "recipients": len(to),
                                                     "status": status})
                return
            except Exception as exc:               # noqa: BLE001 — retried, then logged
                if attempt == 1:
                    await asyncio.sleep(5)
                    continue
                log.error("Notification failed: %s", mailer.friendly_error(exc),
                          extra={"run_id": run_id})
    except Exception as exc:                       # noqa: BLE001 — must never break a run
        log.error("Notification dispatch error: %s", exc, extra={"run_id": run_id})


# ── Weekly digest ────────────────────────────────────────────────
def _send_digest(project_id: int) -> None:
    """Called by APScheduler from a worker thread — plain blocking code."""
    try:
        smtp = _smtp_row(decrypt=True)
        if not (smtp.get("enabled") and (smtp.get("host") or "").strip()):
            return
        cfg = _notify_cfg(project_id)
        if not (cfg.get("enabled") and cfg.get("weekly_digest")):
            return
        to = mailer.parse_recipients(cfg.get("recipients") or "")
        if not to:
            return
        conn = get_db()
        row = conn.execute("SELECT name FROM projects WHERE id=?", (project_id,)).fetchone()
        conn.close()
        if not row:
            return
        fake = {"username": "scheduler", "role": "admin"}
        cov = coverage_report(project_id, stale_days=14, user=fake)
        week = (datetime.now() - timedelta(days=6)).strftime("%Y-%m-%d")
        stats = report_stats(project_id, limit=100, date_from=week,
                             date_to=datetime.now().strftime("%Y-%m-%d"), user=fake)
        subject, text, html_body = mailer.digest_email(
            row["name"], project_id, cov, stats["summary"])
        mailer.send_mail(smtp, to, subject, text, html_body)
        log.info("Weekly digest sent", extra={"project_id": project_id, "recipients": len(to)})
    except Exception as exc:                       # noqa: BLE001
        log.error("Weekly digest failed for project %s: %s", project_id,
                  mailer.friendly_error(exc) if isinstance(exc, Exception) else exc)


def _reload_all_jobs() -> None:
    """Re-register every scheduled job.

    reload_schedules() calls remove_all_jobs(), which would silently delete the
    digest and maintenance jobs too. Always go through this so they cannot drift.
    """
    reload_schedules(get_db, _trigger_group_run)
    _reload_digests()
    _reload_sync_jobs()
    _register_maintenance_job()


def _scheduled_tc_sync(project_id: int) -> None:
    """Automatic reconcile. Runs in an APScheduler worker thread — plain blocking
    code, no event loop involved."""
    try:
        _run_tc_sync(project_id, "scheduler")
    except Exception as exc:                           # noqa: BLE001
        log.warning("Scheduled test case sync failed for project %s: %s", project_id, exc)


def _reload_sync_jobs() -> None:
    """Re-register per-project git sync jobs, under their own id prefix."""
    from scheduler import scheduler
    try:
        for job in scheduler.get_jobs():
            if job.id.startswith("tcsync_"):
                scheduler.remove_job(job.id)
        conn = get_db()
        rows = conn.execute(
            "SELECT id, sync_cron FROM projects"
            " WHERE sync_mode='git' AND sync_cron IS NOT NULL AND sync_cron != ''"
        ).fetchall()
        conn.close()
        for r in rows:
            try:
                scheduler.add_job(_scheduled_tc_sync, trigger=build_trigger(r["sync_cron"]),
                                  args=[r["id"]], id=f"tcsync_{r['id']}",
                                  replace_existing=True)
            except Exception as exc:                   # noqa: BLE001
                log.warning("Bad sync cron for project %s: %s", r["id"], exc)
    except Exception as exc:                           # noqa: BLE001 — never block a save
        log.warning("Could not reload sync jobs: %s", exc)


def _run_maintenance_job() -> None:
    """Nightly housekeeping. Called by APScheduler from a worker thread.

    VACUUM is suppressed while anything is executing: it takes an exclusive lock
    for a full file rewrite, and a test finishing during that window would block
    on its status update long enough to matter.
    """
    busy = bool(_active_runs) or bool(_active_procs)
    res = maintenance.run_maintenance(skip_run_ids=set(_active_runs),
                                      allow_vacuum=not busy)
    if res.get("runs", {}).get("runs") or res.get("orphans", {}).get("dirs"):
        audit("scheduler", "maintenance.purge",
              runs=res["runs"]["runs"], items=res["runs"]["items"],
              orphan_dirs=res["orphans"]["dirs"],
              freed_bytes=res["runs"]["freed_bytes"] + res["orphans"]["freed_bytes"])


def _register_maintenance_job() -> None:
    from scheduler import scheduler
    try:
        for job in scheduler.get_jobs():
            if job.id == "brace_maintenance":
                scheduler.remove_job(job.id)
        # The job always runs — the orphan sweep and ANALYZE are useful even with
        # retention off. Only the run purge is gated on BRACE_RETENTION_DAYS.
        scheduler.add_job(_run_maintenance_job,
                          trigger=build_trigger(maintenance.MAINT_CRON),
                          id="brace_maintenance", replace_existing=True)
    except Exception as exc:                       # noqa: BLE001 — never block startup
        log.warning("Could not register the maintenance job (cron %r): %s",
                    maintenance.MAINT_CRON, exc)


def _reload_digests() -> None:
    """Re-register digest jobs. Uses its own id prefix so it cannot collide with
    the suite schedules, which reload_schedules() clears wholesale."""
    from scheduler import scheduler
    try:
        for job in scheduler.get_jobs():
            if job.id.startswith("digest_"):
                scheduler.remove_job(job.id)
        conn = get_db()
        rows = conn.execute(
            "SELECT project_id, digest_cron FROM notify_config"
            " WHERE enabled=1 AND weekly_digest=1").fetchall()
        conn.close()
        for r in rows:
            try:
                scheduler.add_job(_send_digest, trigger=build_trigger(r["digest_cron"]),
                                  args=[r["project_id"]], id=f"digest_{r['project_id']}",
                                  replace_existing=True)
            except Exception as exc:               # noqa: BLE001
                log.warning("Bad digest cron for project %s: %s", r["project_id"], exc)
    except Exception as exc:                       # noqa: BLE001 — never block a save
        log.warning("Could not reload digest jobs: %s", exc)


# ── Run cancel ───────────────────────────────────────────────────
@app.post("/api/runs/{run_id}/cancel")
async def cancel_run(run_id: str, user=Depends(_current_user)):
    conn = get_db()
    tr = _owned_row(conn, "SELECT project_id FROM test_runs WHERE run_id=?",
                    run_id, user, "run", "Run not found")
    # A queued run has no process yet — flag it so it aborts when its slot
    # comes up, instead of starting after the user already cancelled it.
    _cancelled_runs.add(run_id)

    # Kill any active processes for this run
    for key in list(_active_procs.keys()):
        if key.startswith(run_id):
            proc = _active_procs.get(key)
            if proc:
                proc.terminate()
    now = datetime.now().isoformat()
    conn.execute("UPDATE test_runs SET status='cancelled', finished_at=? WHERE run_id=?",
                 (now, run_id))
    conn.execute("UPDATE test_run_items SET status='cancelled', finished_at=?"
                 " WHERE run_id=? AND status IN ('pending','running')", (now, run_id))
    conn.commit()
    conn.close()
    _active_runs.pop(run_id, None)
    _metrics["runs_cancelled"] += 1
    log.info("Run cancelled", extra={"run_id": run_id, "user": user["username"]})
    audit(user, "run.cancel", project_id=tr["project_id"], target=run_id)
    _publish(run_id, "done", {"status": "cancelled"})
    return {"cancelled": run_id}


# ── Run listing & detail ─────────────────────────────────────────
@app.get("/api/projects/{project_id}/runs")
def list_runs(project_id: int, user=Depends(_proj_viewer)):
    conn = get_db()
    rows = conn.execute("""
        SELECT tr.*, tg.name AS group_name
        FROM test_runs tr
        LEFT JOIN test_groups tg ON tr.group_id = tg.id
        WHERE tr.project_id=? ORDER BY tr.started_at DESC, tr.id DESC LIMIT 100
    """, (project_id,)).fetchall()
    conn.close()
    result = rows_to_list(rows)
    # Merge in-memory live state
    for r in result:
        live = _active_runs.get(r["run_id"])
        if live:
            r["status"] = live["status"]
            r["passed"] = live["passed"]
            r["failed"] = live["failed"]
    return result


@app.get("/api/projects/{project_id}/coverage")
def coverage_report(project_id: int, stale_days: int = 14, user=Depends(_proj_viewer)):
    """Automation coverage — an inventory view, not an execution view.

    The Reports dashboard answers "how did the runs in this window go?". This
    answers a different question: "of everything we have, how much is actually
    being tested?" — which is what tells you whether a 100% pass rate is
    meaningful or just the same three cases passing over and over.

    Deliberately NOT date-filtered: coverage is about the current state of the
    inventory. `stale_days` only decides how old a last run must be to count as
    stale.
    """
    conn = get_db()
    tcs = rows_to_list(conn.execute(
        "SELECT id, tc_code, name, suite_path, tags, last_run_status, last_run_at"
        " FROM test_cases WHERE project_id=? ORDER BY tc_code", (project_id,)).fetchall())

    # Suite membership, for the per-suite breakdown. A case can be in several
    # suites; one not in any is reported under "(no suite)" because those are
    # exactly the cases that never run as part of a scheduled pack.
    memb: dict = {}
    for r in conn.execute("""
            SELECT gtc.test_case_id AS tid, tg.name AS gname
            FROM group_test_cases gtc JOIN test_groups tg ON tg.id = gtc.group_id
            WHERE tg.project_id=?""", (project_id,)).fetchall():
        memb.setdefault(r["tid"], []).append(r["gname"])
    conn.close()

    # Scripts on disk, using the same rule as the suite-path picker so the two
    # can never disagree about what counts as a runnable script.
    suites_dir = _project_suites(project_id)
    scripts = set()
    if suites_dir.exists():
        for p in suites_dir.rglob("*.robot"):
            rel = str(p.relative_to(suites_dir)).replace("\\", "/")
            if any(part == "testcases" for part in rel.lower().split("/")[:-1]):
                scripts.add(rel)

    linked = {(t["suite_path"] or "").strip() for t in tcs if (t["suite_path"] or "").strip()}
    orphan_scripts = sorted(scripts - linked)          # on disk, no test case
    missing_scripts = sorted(p for p in linked if p not in scripts)  # case points nowhere

    cutoff = (datetime.now() - timedelta(days=max(1, stale_days))).isoformat(timespec="seconds")

    def bucket(t: dict) -> str:
        if not t["last_run_status"]:
            return "never"
        if t["last_run_status"] == "passed":
            return "passed"
        return "failed"

    total = len(tcs)
    counts = {"passed": 0, "failed": 0, "never": 0}
    stale, never_list, failing_list, stale_list = 0, [], [], []
    for t in tcs:
        b = bucket(t)
        counts[b] += 1
        brief = {"tc_code": t["tc_code"], "name": t["name"],
                 "suite_path": t["suite_path"], "last_run_at": t["last_run_at"],
                 "suites": memb.get(t["id"], [])}
        if b == "never":
            never_list.append(brief)
        else:
            if b == "failed":
                failing_list.append(brief)
            # Compare on the date/time prefix: last_run_at is ISO with 'T', and
            # so is the cutoff, so this is a like-for-like string compare.
            if (t["last_run_at"] or "") < cutoff:
                stale += 1
                stale_list.append(brief)

    executed = total - counts["never"]
    pct = lambda n, d: round(n / d * 100, 1) if d else 0.0

    # Per-suite coverage: a suite of 40 cases where only 4 ever ran is the
    # thing worth seeing, and the project-level number hides it.
    by_suite: dict = {}
    for t in tcs:
        for g in (memb.get(t["id"]) or ["(no suite)"]):
            s = by_suite.setdefault(g, {"suite": g, "total": 0, "passed": 0,
                                        "failed": 0, "never": 0})
            s["total"] += 1
            s[bucket(t)] += 1
    for s in by_suite.values():
        s["executed"] = s["total"] - s["never"]
        s["coverage_pct"] = pct(s["executed"], s["total"])
        s["pass_pct"] = pct(s["passed"], s["executed"])

    return {
        "stale_days": stale_days,
        "scripts": {
            "on_disk":  len(scripts),
            "onboarded": len(scripts & linked),
            "orphan":   len(orphan_scripts),
            "onboarded_pct": pct(len(scripts & linked), len(scripts)),
            "orphan_list":  orphan_scripts[:100],
            "missing_list": missing_scripts[:100],
        },
        "test_cases": {
            "total":        total,
            "executed":     executed,
            "never":        counts["never"],
            "passed":       counts["passed"],
            "failed":       counts["failed"],
            "stale":        stale,
            "coverage_pct": pct(executed, total),      # ever ran at least once
            "passed_pct":   pct(counts["passed"], total),
            "failed_pct":   pct(counts["failed"], total),
            "never_pct":    pct(counts["never"], total),
            # Of the cases that have run, how many are currently green. Not the
            # same as the Reports pass rate, which is per execution.
            "health_pct":   pct(counts["passed"], executed),
        },
        "by_suite":  sorted(by_suite.values(), key=lambda s: (s["coverage_pct"], -s["total"])),
        "never_run": never_list[:200],
        "failing":   failing_list[:200],
        "stale_list": stale_list[:200],
    }


@app.get("/api/projects/{project_id}/report-stats")
def report_stats(project_id: int, limit: int = 30,
                  date_from: Optional[str] = None, date_to: Optional[str] = None,
                  user=Depends(_proj_viewer)):
    """Aggregated analytics for the Reports dashboard.

    date_from/date_to are inclusive calendar dates (YYYY-MM-DD). Default: today only.
    """
    if not date_from and not date_to:
        today = datetime.now().strftime("%Y-%m-%d")
        date_from = date_to = today
    date_from = date_from or "0001-01-01"
    date_to   = date_to   or "9999-12-31"
    # Compare the calendar date only. started_at is written by Python as
    # '2026-08-13T17:28:15' (ISO, 'T' separator) while older rows written by
    # SQLite's datetime('now') use a space. Comparing whole timestamps against
    # a '<date> 23:59:59' bound silently excluded every ISO row, because
    # ord('T') > ord(' ') puts it above the upper bound — the Reports page
    # showed zero runs for a day that clearly had them.
    range_start = date_from[:10]
    range_end   = date_to[:10]

    conn = get_db()

    runs = rows_to_list(conn.execute("""
        SELECT tr.run_id, tr.run_name, tr.status, tr.total, tr.passed, tr.failed,
               tr.started_at, tr.finished_at, tr.triggered_by, tg.name AS group_name
        FROM test_runs tr LEFT JOIN test_groups tg ON tr.group_id = tg.id
        WHERE tr.project_id=? AND substr(tr.started_at,1,10) BETWEEN ? AND ?
        ORDER BY tr.started_at DESC, tr.id DESC LIMIT ?
    """, (project_id, range_start, range_end, limit)).fetchall())

    # Per-TC aggregation across runs within the date window
    tc_rows = rows_to_list(conn.execute("""
        SELECT tri.tc_code, tri.tc_name,
               COUNT(*)                                            AS runs,
               SUM(CASE WHEN tri.status='passed' THEN 1 ELSE 0 END) AS passed,
               SUM(CASE WHEN tri.status='failed' THEN 1 ELSE 0 END) AS failed,
               MAX(tri.finished_at)                                AS last_run
        FROM test_run_items tri
        JOIN test_runs tr ON tri.run_id = tr.run_id
        WHERE tr.project_id=? AND substr(tr.started_at,1,10) BETWEEN ? AND ?
        GROUP BY tri.tc_code, tri.tc_name
        ORDER BY failed DESC, runs DESC
    """, (project_id, range_start, range_end)).fetchall())

    # Totals within the same date window
    agg = conn.execute("""
        SELECT COUNT(*) AS n_runs,
               COALESCE(SUM(total),0)  AS tot,
               COALESCE(SUM(passed),0) AS pass,
               COALESCE(SUM(failed),0) AS fail
        FROM test_runs WHERE project_id=? AND substr(started_at,1,10) BETWEEN ? AND ?
    """, (project_id, range_start, range_end)).fetchone()
    conn.close()

    def _dur(a: Optional[str], b: Optional[str]) -> Optional[float]:
        if not a or not b:
            return None
        try:
            return (datetime.fromisoformat(b) - datetime.fromisoformat(a)).total_seconds()
        except ValueError:
            return None

    durations = []
    for r in runs:
        d = _dur(r.get("started_at"), r.get("finished_at"))
        r["duration_sec"] = d
        if d is not None:
            durations.append(d)

    # Flaky = a TC that has both passed and failed at least once
    flaky = [t for t in tc_rows if t["passed"] > 0 and t["failed"] > 0]
    for t in tc_rows:
        t["pass_rate"] = round(t["passed"] / t["runs"] * 100) if t["runs"] else 0

    total_tc = agg["tot"] or 0
    return {
        "date_from": date_from,
        "date_to":   date_to,
        "summary": {
            "total_runs":   agg["n_runs"] or 0,
            "total_tests":  total_tc,
            "total_passed": agg["pass"] or 0,
            "total_failed": agg["fail"] or 0,
            "pass_rate":    round((agg["pass"] or 0) / total_tc * 100, 1) if total_tc else 0.0,
            "avg_duration": round(sum(durations) / len(durations), 1) if durations else None,
            "flaky_count":  len(flaky),
        },
        # oldest → newest so the trend chart reads left-to-right
        "trend":        list(reversed(runs)),
        "top_failing":  [t for t in tc_rows if t["failed"] > 0][:10],
        "flaky":        flaky[:10],
        "tc_stats":     tc_rows,
    }


def _run_or_404(conn, run_id: str, user):
    tr = conn.execute("""
        SELECT tr.*, tg.name AS group_name
        FROM test_runs tr LEFT JOIN test_groups tg ON tr.group_id = tg.id
        WHERE tr.run_id=?
    """, (run_id,)).fetchone()
    if not tr:
        conn.close()
        raise HTTPException(404, "Run not found")
    if not _get_project_role(tr["project_id"], user["username"], user["role"]):
        conn.close()
        raise HTTPException(403)
    return tr


def _item_files(run_dir, rf_run_id):
    """(has_log, has_report) for one item. Two stat() calls — only ever done for
    the page of items actually being returned, never for the whole run."""
    if not rf_run_id:
        return False, False
    d = run_dir / rf_run_id
    return (d / "log.html").exists(), (d / "report.html").exists()


@app.get("/api/runs/{run_id}")
def get_run(run_id: str, include_items: bool = False, user=Depends(_current_user)):
    """Run summary + per-status counts.

    Items are NOT included by default. A 1200-case run serialises to ~525 KB
    with them, and the detail view polls this every 3 s — that was 175 KB/s per
    viewer just to redraw a progress bar. The UI pages through
    /api/runs/{id}/items instead. include_items=1 restores the old shape for
    any external caller that still wants everything in one response.
    """
    conn = get_db()
    tr = _run_or_404(conn, run_id, user)
    counts = {r["status"]: r["n"] for r in conn.execute(
        "SELECT status, COUNT(*) AS n FROM test_run_items WHERE run_id=? GROUP BY status",
        (run_id,)).fetchall()}
    items = conn.execute(
        "SELECT * FROM test_run_items WHERE run_id=? ORDER BY id", (run_id,)
    ).fetchall() if include_items else []
    conn.close()

    d = dict(tr)
    live = _active_runs.get(run_id)
    if live:
        d["status"] = live["status"]
        d["passed"] = live["passed"]
        d["failed"] = live["failed"]

    run_dir = _project_results(tr["project_id"]) / run_id
    d["status_counts"] = counts
    d["item_count"]    = sum(counts.values())
    if include_items:
        item_list = []
        for it in items:
            i = dict(it)
            i["has_log"], i["has_report"] = _item_files(run_dir, it["rf_run_id"])
            item_list.append(i)
        d["items"] = item_list
    d["has_combined_report"] = (run_dir / "combined" / "report.html").exists()
    d["has_combined_log"]    = (run_dir / "combined" / "log.html").exists()
    return d


# Columns the list view needs. fail_detail is deliberately absent — it is the
# bulk of the payload (stack traces, page source) and is only ever read for the
# one row a user expands, which /items/{item_id} serves.
_ITEM_LIST_COLS = ("id, test_case_id, tc_code, tc_name, status, rf_run_id, "
                   "started_at, finished_at, fail_summary")


@app.get("/api/runs/{run_id}/items")
def get_run_items(run_id: str,
                  status: str = "",
                  q: str = "",
                  offset: int = 0,
                  limit: int = 50,
                  user=Depends(_current_user)):
    """One page of a run's test cases, filtered by status and free text."""
    limit  = max(1, min(500, limit))
    offset = max(0, offset)
    conn = get_db()
    tr = _run_or_404(conn, run_id, user)

    where, params = ["run_id=?"], [run_id]
    if status:
        marks = ",".join("?" for _ in status.split(","))
        where.append(f"status IN ({marks})")
        params += [s.strip() for s in status.split(",")]
    if q.strip():
        where.append("(tc_code LIKE ? OR tc_name LIKE ? OR fail_summary LIKE ?)")
        params += [f"%{q.strip()}%"] * 3
    clause = " AND ".join(where)

    total = conn.execute(
        f"SELECT COUNT(*) AS n FROM test_run_items WHERE {clause}", params
    ).fetchone()["n"]
    rows = conn.execute(
        f"SELECT {_ITEM_LIST_COLS} FROM test_run_items WHERE {clause} "
        f"ORDER BY id LIMIT ? OFFSET ?", params + [limit, offset]
    ).fetchall()
    conn.close()

    run_dir = _project_results(tr["project_id"]) / run_id
    out = []
    for r in rows:
        i = dict(r)
        i["has_log"], i["has_report"] = _item_files(run_dir, r["rf_run_id"])
        out.append(i)
    return {"total": total, "offset": offset, "limit": limit, "items": out}


@app.get("/api/runs/{run_id}/items/{item_id}")
def get_run_item(run_id: str, item_id: int, user=Depends(_current_user)):
    """Full detail for one test case, including the failure text and screenshot."""
    conn = get_db()
    tr = _run_or_404(conn, run_id, user)
    row = conn.execute(
        "SELECT * FROM test_run_items WHERE run_id=? AND id=?", (run_id, item_id)
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Test case not found in this run")
    i = dict(row)
    i["has_log"], i["has_report"] = _item_files(
        _project_results(tr["project_id"]) / run_id, row["rf_run_id"])
    return i


@app.get("/api/runs/{run_id}/events")
async def stream_run_events(run_id: str, request: Request, token: str = "",
                            user=Depends(oauth2_scheme)):
    """Server-sent events for one run: status, per-case verdicts, completion.

    EventSource cannot set an Authorization header, so the token also comes in
    as a query parameter — the same accommodation the results route already
    makes for <img> and <iframe>.

    Replaces polling while a run is live. The 3s poll re-queried the whole run
    for every watcher; this pushes ~100 bytes per case as it finishes.
    """
    raw = token or user
    if not raw:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(raw, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        who = {"username": payload["sub"], "role": payload.get("role", "user")}
    except JWTError:
        raise HTTPException(401, "Invalid or expired token")

    conn = get_db()
    tr = conn.execute("SELECT project_id, status, passed, failed, total"
                      " FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    if not tr:
        conn.close()
        raise HTTPException(404, "Run not found")
    if not _get_project_role(tr["project_id"], who["username"], who["role"]):
        conn.close()
        raise HTTPException(403)
    counts = {r["status"]: r["n"] for r in conn.execute(
        "SELECT status, COUNT(*) AS n FROM test_run_items WHERE run_id=? GROUP BY status",
        (run_id,)).fetchall()}
    conn.close()

    subs = _run_subs.setdefault(run_id, set())
    if len(subs) >= SSE_MAX_SUBS:
        # Not an error: the client falls back to polling, which still works.
        raise HTTPException(503, "Too many live watchers for this run")
    queue: asyncio.Queue = asyncio.Queue(maxsize=1000)
    subs.add(queue)

    def _sse(event: str, data: dict) -> str:
        import json as _json
        return f"event: {event}\ndata: {_json.dumps(data, default=str)}\n\n"

    async def generate():
        try:
            live = _active_runs.get(run_id)
            yield _sse("summary", {
                "status": (live or tr)["status"], "passed": (live or tr)["passed"],
                "failed": (live or tr)["failed"], "total": tr["total"],
                "status_counts": counts})
            # A run that already finished gets one summary and an immediate
            # 'done' — no connection is left hanging for a run nothing will
            # ever publish to again.
            if tr["status"] not in ("running", "queued"):
                yield _sse("done", {"status": tr["status"]})
                return
            while True:
                if await request.is_disconnected():
                    return
                try:
                    event, data = await asyncio.wait_for(queue.get(),
                                                         timeout=SSE_HEARTBEAT_SEC)
                except asyncio.TimeoutError:
                    yield ": keep-alive\n\n"     # comment frame; ignored by EventSource
                    continue
                yield _sse(event, data)
                if event == "done":
                    return
        finally:
            # Always unsubscribe — a leaked queue would keep _publish fanning
            # out to a client that disconnected hours ago.
            subs.discard(queue)
            if not subs:
                _run_subs.pop(run_id, None)

    return StreamingResponse(generate(), media_type="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "Connection":    "keep-alive",
        # nginx/OCP router buffering would defeat the entire point of streaming.
        "X-Accel-Buffering": "no",
    })


@app.get("/api/runs/{run_id}/log")
async def stream_run_log(run_id: str, user=Depends(_current_user)):
    """Stream console logs from all items in a run."""
    conn = get_db()
    tr = conn.execute("SELECT project_id FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    if not tr:
        conn.close()
        raise HTTPException(404)
    role = _get_project_role(tr["project_id"], user["username"], user["role"])
    if not role:
        conn.close()
        raise HTTPException(403)
    project_id = tr["project_id"]
    conn.close()

    run_dir = _project_results(project_id) / run_id

    async def generate():
        pos = 0
        elapsed = 0
        # Collect logs from all item subdirs
        while elapsed < SSE_TIMEOUT:
            # Find all console.log files in run_dir
            log_files = sorted(run_dir.rglob("console.log")) if run_dir.exists() else []
            full_text = ""
            for lf in log_files:
                try:
                    full_text += f"\n--- {lf.parent.name} ---\n"
                    full_text += lf.read_text(errors="replace")
                except Exception:
                    pass
            if len(full_text) > pos:
                chunk = full_text[pos:]
                pos   = len(full_text)
                for line in chunk.splitlines():
                    yield f"data: {line}\n\n"

            live = _active_runs.get(run_id)
            if live and live.get("status") in ("passed", "failed", "cancelled"):
                yield "data: __DONE__\n\n"
                break
            if not live:
                conn2 = get_db()
                row = conn2.execute("SELECT status FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
                conn2.close()
                if row and row["status"] in ("passed", "failed", "cancelled"):
                    yield "data: __DONE__\n\n"
                    break

            await asyncio.sleep(2)
            elapsed += 2
        else:
            yield "data: __TIMEOUT__\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── Report file serving ──────────────────────────────────────────
@app.get("/results/{project_id}/{run_id}/{subpath:path}")
async def serve_result(project_id: int, run_id: str, subpath: str,
                       request: Request, token: Optional[str] = None):
    # Accept token from query param (iframe) or Authorization header (API calls)
    raw = token or (request.headers.get("authorization", "").removeprefix("Bearer ").strip() or None)
    if not raw:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(raw, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = {"username": payload["sub"], "role": payload.get("role", "user")}
    except JWTError:
        raise HTTPException(401, "Invalid token")
    role = _get_project_role(project_id, user["username"], user["role"])
    if not role:
        raise HTTPException(403)
    # run_id and subpath come straight from the URL — confine them to the
    # project's results dir before touching the filesystem.
    base = _project_results(project_id).resolve()
    path = (base / run_id / subpath).resolve()
    if not _contained(base, path):
        raise HTTPException(400, "Path traversal denied")
    if not path.is_file():
        raise HTTPException(404, "File not found")
    return FileResponse(str(path))


# ══════════════════════════════════════════════════════════════════
# AI DEBUG ASSISTANT
# ══════════════════════════════════════════════════════════════════

MAX_CONSOLE_CHARS = 8000
MAX_SOURCE_CHARS  = 12000
MAX_RESOURCE_FILES = 4


def _hush_insecure_warning(verify: bool) -> None:
    """Silence urllib3's per-request warning when verification is off by choice."""
    if verify:
        return
    try:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    except Exception:                                # noqa: BLE001 — best-effort only
        pass


def _get_ai_config() -> dict:
    conn = get_db()
    row = conn.execute("SELECT * FROM ai_config WHERE id=1").fetchone()
    conn.close()
    if not row:
        return {"enabled": False, "api_base": "", "api_key": "", "model": "", "verify_ssl": True}
    d = dict(row)
    d["enabled"]    = bool(d.get("enabled"))
    d["verify_ssl"] = bool(d.get("verify_ssl", 1))
    d["api_key"]    = decrypt_token(d.get("api_key") or "")
    return d


def _parse_output_xml(path: Path) -> list:
    """Extract failing tests + the keyword chain that failed from RF output.xml."""
    import xml.etree.ElementTree as ET
    if not path.exists():
        return []
    try:
        tree = ET.parse(str(path))
    except ET.ParseError:
        return []
    root = tree.getroot()
    failures = []

    def walk_kws(node, trail):
        """Depth-first through <kw>, recording the deepest FAIL trail."""
        found = []
        for kw in node.findall("kw"):
            st = kw.find("status")
            if st is None or st.get("status") != "FAIL":
                continue
            name = kw.get("name") or ""
            lib  = kw.get("library") or kw.get("owner") or ""
            label = f"{lib}.{name}" if lib else name
            args = [a.text or "" for a in kw.findall("arguments/arg")]
            new_trail = trail + [{"keyword": label, "args": args}]
            msgs = [(m.text or "").strip() for m in kw.findall("msg")
                    if m.get("level") in ("FAIL", "ERROR")]
            deeper = walk_kws(kw, new_trail)
            if deeper:
                found.extend(deeper)
            else:
                found.append({"trail": new_trail, "messages": msgs,
                              "status_text": (st.text or "").strip()})
        return found

    for test in root.iter("test"):
        st = test.find("status")
        if st is None or st.get("status") != "FAIL":
            continue
        chains = walk_kws(test, [])
        # Setup/teardown failures live outside the <kw> chain
        failures.append({
            "test_name": test.get("name") or "",
            "message":   (st.text or "").strip(),
            "chains":    chains[:3],
        })
    return failures


def _read_capped(path: Path, cap: int, tail: bool = False) -> str:
    try:
        # Explicit UTF-8 — don't inherit whatever locale the container happens to have
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return ""
    if len(text) <= cap:
        return text
    return ("…(truncated)…\n" + text[-cap:]) if tail else (text[:cap] + "\n…(truncated)…")


def _collect_resources(robot_src: str, suites_dir: Path, suite_path: str) -> list:
    """Pull the Resource/Variables files a suite imports, so the model sees the keywords."""
    import re
    out = []
    base = (suites_dir / suite_path).parent if suite_path else suites_dir
    pattern = re.compile(r"^(?:Resource|Variables)\s+(.+?)\s*$", re.MULTILINE | re.IGNORECASE)
    for m in pattern.finditer(robot_src):
        raw = m.group(1).strip()
        if "$" in raw:      # variable-based path — can't resolve statically
            continue
        cand = (base / raw).resolve()
        try:
            cand.relative_to(suites_dir.resolve())
        except ValueError:
            continue
        if cand.exists() and cand.is_file():
            out.append({"path": str(cand.relative_to(suites_dir.resolve())).replace("\\", "/"),
                        "content": _read_capped(cand, MAX_SOURCE_CHARS // 2)})
        if len(out) >= MAX_RESOURCE_FILES:
            break
    return out


def _build_debug_context(project_id: int, run_id: str,
                         rf_run_id: Optional[str], suite_path: Optional[str]) -> dict:
    suites_dir = _project_suites(project_id)
    result_dir = _project_results(project_id) / run_id
    if rf_run_id:
        result_dir = result_dir / rf_run_id

    console  = _read_capped(result_dir / "console.log", MAX_CONSOLE_CHARS, tail=True)
    failures = _parse_output_xml(result_dir / "output.xml")

    # Fall back to the run item's suite path when the caller didn't supply one
    if not suite_path and rf_run_id:
        conn = get_db()
        row = conn.execute("""
            SELECT tc.suite_path FROM test_run_items tri
            LEFT JOIN test_cases tc ON tri.test_case_id = tc.id
            WHERE tri.run_id=? AND tri.rf_run_id=?
        """, (run_id, rf_run_id)).fetchone()
        conn.close()
        if row:
            suite_path = row["suite_path"]

    robot_src, resources = "", []
    if suite_path:
        f = suites_dir / suite_path
        if f.exists() and f.is_file():
            robot_src = _read_capped(f, MAX_SOURCE_CHARS)
            resources = _collect_resources(robot_src, suites_dir, suite_path)

    return {
        "project_id": project_id,
        "run_id":     run_id,
        "rf_run_id":  rf_run_id,
        "suite_path": suite_path,
        "console":    console,
        "failures":   failures,
        "robot_src":  robot_src,
        "resources":  resources,
    }


def _render_debug_prompt(ctx: dict) -> str:
    """Self-contained prompt — usable verbatim in any external chat when air-gapped."""
    p = []
    p.append("You are a Robot Framework test-automation debugging expert. "
             "A test has failed. Diagnose the root cause and give a concrete fix.\n")
    p.append("## Environment")
    p.append("- Robot Framework test running headless in a Linux container (Chrome + ChromeDriver via Xvfb)")
    p.append(f"- Suite file: `{ctx.get('suite_path') or 'unknown'}`")
    p.append(f"- Run ID: `{ctx.get('run_id')}`\n")

    fails = ctx.get("failures") or []
    if fails:
        p.append("## Failures (parsed from output.xml)")
        for f in fails:
            p.append(f"\n### Test: {f['test_name']}")
            p.append(f"**Status message:** {f['message'] or '(none)'}")
            for i, ch in enumerate(f.get("chains") or [], 1):
                trail = ch.get("trail") or []
                if trail:
                    p.append(f"\n**Failing keyword chain {i}:**")
                    for depth, kw in enumerate(trail):
                        args = ", ".join(kw["args"]) if kw["args"] else ""
                        p.append(f"{'  ' * depth}- `{kw['keyword']}`" + (f"  args: `{args}`" if args else ""))
                for msg in (ch.get("messages") or []):
                    p.append(f"\n**Error message:**\n```\n{msg}\n```")
        p.append("")

    if ctx.get("console"):
        p.append("## Console output (tail)")
        p.append("```\n" + ctx["console"] + "\n```\n")

    if ctx.get("robot_src"):
        p.append(f"## Test suite source — `{ctx.get('suite_path')}`")
        p.append("```robotframework\n" + ctx["robot_src"] + "\n```\n")

    for r in (ctx.get("resources") or []):
        p.append(f"## Imported resource — `{r['path']}`")
        p.append("```robotframework\n" + r["content"] + "\n```\n")

    p.append("## What I need from you")
    p.append("1. **Root cause** — what actually failed and why (be specific about the keyword and data).")
    p.append("2. **Category** — one of: locator/selector issue, timing/synchronisation, test data, "
             "environment/config, application defect, or script logic bug.")
    p.append("3. **Fix** — the exact change, as a code diff or replacement snippet against the source above.")
    p.append("4. **Verification** — how to confirm the fix worked.")
    p.append("\nIf the evidence is insufficient for a confident diagnosis, say what additional "
             "log or screenshot would settle it, rather than guessing.")
    return "\n".join(p)


class AIDebugReq(BaseModel):
    run_id:     str
    rf_run_id:  Optional[str] = None
    suite_path: Optional[str] = None


@app.post("/api/projects/{project_id}/ai-debug/prompt")
def ai_debug_prompt(project_id: int, req: AIDebugReq, user=Depends(_proj_viewer)):
    """Build the debug prompt. Always available — this is the air-gapped path."""
    ctx = _build_debug_context(project_id, req.run_id, req.rf_run_id, req.suite_path)
    cfg = _get_ai_config()
    return {
        "prompt":       _render_debug_prompt(ctx),
        "ai_available": bool(cfg["enabled"] and cfg["api_key"]),
        "model":        cfg["model"] if cfg["enabled"] else None,
        "has_failures": bool(ctx["failures"]),
        "suite_path":   ctx["suite_path"],
    }


@app.post("/api/projects/{project_id}/ai-debug/analyze")
def ai_debug_analyze(project_id: int, req: AIDebugReq, user=Depends(_proj_viewer)):
    """Stream a diagnosis from an OpenAI-compatible endpoint (OpenRouter, vLLM, LiteLLM…)."""
    cfg = _get_ai_config()
    if not cfg["enabled"] or not cfg["api_key"]:
        raise HTTPException(400, "AI assistant not configured. Use the prompt option instead.")

    ctx    = _build_debug_context(project_id, req.run_id, req.rf_run_id, req.suite_path)
    prompt = _render_debug_prompt(ctx)

    def stream():
        import json
        import requests
        url = cfg["api_base"].rstrip("/") + "/chat/completions"
        _hush_insecure_warning(cfg["verify_ssl"])
        try:
            resp = requests.post(
                url,
                headers={"Authorization": f"Bearer {cfg['api_key']}",
                         "Content-Type": "application/json"},
                json={"model": cfg["model"],
                      "messages": [{"role": "user", "content": prompt}],
                      "stream": True},
                stream=True, timeout=180, verify=cfg["verify_ssl"],
            )
            # SSE responses carry no charset param, so requests falls back to
            # ISO-8859-1 and mangles every non-ASCII byte. Force UTF-8.
            resp.encoding = "utf-8"
            if resp.status_code != 200:
                yield f"\n[AI ERROR {resp.status_code}] {resp.text[:600]}\n"
                return
            for line in resp.iter_lines(decode_unicode=True):
                if not line or not line.startswith("data: "):
                    continue
                data = line[6:].strip()
                if data == "[DONE]":
                    break
                try:
                    delta = json.loads(data)["choices"][0].get("delta", {})
                except (ValueError, KeyError, IndexError):
                    continue
                if delta.get("content"):
                    yield delta["content"]
        except Exception as exc:                      # noqa: BLE001 — surfaced to the UI
            yield f"\n[AI ERROR] {type(exc).__name__}: {exc}\n"

    return StreamingResponse(stream(), media_type="text/plain")


class AIConfigReq(BaseModel):
    enabled:    bool = False
    api_base:   str = "https://openrouter.ai/api/v1"
    api_key:    Optional[str] = None     # omit / blank to keep the stored key
    model:      str = "anthropic/claude-sonnet-4"
    verify_ssl: bool = True              # off for internal routes with self-signed certs


# ══════════════════════════════════════════════════════════════════
# EMAIL NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════

class SmtpConfigReq(BaseModel):
    enabled:     Optional[bool] = None
    host:        Optional[str]  = None
    port:        Optional[int]  = None
    security:    Optional[str]  = None      # starttls | ssl | none
    username:    Optional[str]  = None
    password:    Optional[str]  = None      # blank on save keeps the stored one
    from_addr:   Optional[str]  = None
    from_name:   Optional[str]  = None
    verify_ssl:  Optional[bool] = None
    timeout_sec: Optional[int]  = None


def _smtp_row(decrypt: bool = False) -> dict:
    conn = get_db()
    row = conn.execute("SELECT * FROM smtp_config WHERE id=1").fetchone()
    conn.close()
    d = dict(row) if row else {}
    d["password"] = decrypt_token(d.get("password") or "") if decrypt else ""
    return d


@app.get("/api/admin/smtp-presets")
def smtp_presets(user=Depends(_require_sys_admin)):
    """Host/port/security for common providers. Never credentials."""
    return mailer.SMTP_PRESETS


@app.get("/api/admin/smtp-config")
def get_smtp_config(user=Depends(_require_sys_admin)):
    d = _smtp_row()
    # The password itself is never returned; the UI only needs to know whether
    # one is stored so it can say "leave blank to keep".
    stored = _smtp_row(decrypt=True).get("password")
    d.pop("password", None)
    d["has_password"] = bool(stored)
    d["public_url"] = mailer.PUBLIC_URL or ""
    return d


@app.put("/api/admin/smtp-config")
def put_smtp_config(req: SmtpConfigReq, user=Depends(_require_sys_admin)):
    cur = _smtp_row(decrypt=True)
    sec = (req.security or cur.get("security") or "starttls").lower()
    if sec not in ("starttls", "ssl", "none"):
        raise HTTPException(400, "security must be starttls, ssl or none")
    # An empty password means "keep what is stored" — otherwise every save from
    # a UI that cannot show the password would wipe it.
    pwd = (cur.get("password") if req.password in (None, "")
           else mailer.normalise_password(req.password))
    conn = get_db()
    conn.execute("""
        UPDATE smtp_config SET enabled=?, host=?, port=?, security=?, username=?,
               password=?, from_addr=?, from_name=?, verify_ssl=?, timeout_sec=?,
               updated_at=? WHERE id=1""",
        (1 if (cur.get("enabled") if req.enabled is None else req.enabled) else 0,
         (req.host if req.host is not None else cur.get("host")) or "",
         int(req.port if req.port is not None else (cur.get("port") or 587)),
         sec,
         (req.username if req.username is not None else cur.get("username")) or "",
         encrypt_token(pwd or ""),
         (req.from_addr if req.from_addr is not None else cur.get("from_addr")) or "",
         (req.from_name if req.from_name is not None else cur.get("from_name")) or "BRACE",
         1 if (cur.get("verify_ssl", 1) if req.verify_ssl is None else req.verify_ssl) else 0,
         max(5, int(req.timeout_sec if req.timeout_sec is not None else (cur.get("timeout_sec") or 20))),
         datetime.now().isoformat(timespec="seconds")))
    conn.commit()
    conn.close()
    log.info("SMTP config updated", extra={"user": user["username"]})
    audit(user, "smtp.update", target=req.host or cur.get("host"),
          enabled=req.enabled, security=sec,
          # The value never reaches the log — only whether it was replaced.
          password=bool(req.password))
    return get_smtp_config(user)


class SmtpTestReq(BaseModel):
    to: str


@app.post("/api/admin/smtp-config/test")
async def test_smtp_config(req: SmtpTestReq, user=Depends(_require_sys_admin)):
    """Send a real email. Reports the actual error, which is the whole point —
    SMTP misconfiguration is otherwise invisible until an alert silently fails.
    """
    to = mailer.parse_recipients(req.to)
    if not to:
        raise HTTPException(400, "Enter a valid destination address")
    cfg = _smtp_row(decrypt=True)
    if not (cfg.get("host") or "").strip():
        raise HTTPException(400, "Configure the SMTP host first")
    text = (f"This is a test message from BRACE.\n\n"
            f"If you received it, notifications are correctly configured.\n"
            f"Sent {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} "
            f"from pod {POD_NAME}.\n")
    html_body = ("<div style=\"font-family:'Segoe UI',Arial,sans-serif;padding:20px\">"
                 "<h2 style='color:#0F3278;margin:0 0 8px'>BRACE test email</h2>"
                 "<p style='font-size:14px;color:#1a2340'>If you received this, "
                 "notifications are correctly configured.</p>"
                 f"<p style='font-size:12px;color:#6b7a99'>Sent "
                 f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} from pod "
                 f"{POD_NAME}.</p></div>")
    try:
        await asyncio.to_thread(mailer.send_mail, cfg, to,
                                "BRACE test email", text, html_body)
    except Exception as exc:                       # noqa: BLE001 — reported to the user
        log.warning("SMTP test failed: %s", type(exc).__name__)
        raise HTTPException(400, mailer.friendly_error(exc))
    audit(user, "smtp.test", recipients=len(to))
    return {"ok": True, "sent_to": to}


# ── Per-project notification settings ────────────────────────────
class NotifyConfigReq(BaseModel):
    enabled:             Optional[bool] = None
    on_run_failed:       Optional[bool] = None
    on_scheduled_failed: Optional[bool] = None
    on_run_passed:       Optional[bool] = None
    only_on_change:      Optional[bool] = None
    notify_triggerer:    Optional[bool] = None
    weekly_digest:       Optional[bool] = None
    digest_cron:         Optional[str]  = None
    recipients:          Optional[str]  = None


_NOTIFY_DEFAULTS = {"enabled": 0, "on_run_failed": 1, "on_scheduled_failed": 1,
                    "on_run_passed": 0, "only_on_change": 1, "notify_triggerer": 0,
                    "weekly_digest": 0, "digest_cron": "0 8 * * 1", "recipients": ""}


def _notify_cfg(project_id: int) -> dict:
    conn = get_db()
    row = conn.execute("SELECT * FROM notify_config WHERE project_id=?",
                       (project_id,)).fetchone()
    conn.close()
    d = dict(_NOTIFY_DEFAULTS)
    d["project_id"] = project_id
    if row:
        d.update({k: v for k, v in dict(row).items() if v is not None})
    return d


@app.get("/api/projects/{project_id}/notify-config")
def get_notify_config(project_id: int, user=Depends(_proj_admin)):
    d = _notify_cfg(project_id)
    d["recipient_list"] = mailer.parse_recipients(d.get("recipients") or "")
    # Notifications are useless without a working transport; say so explicitly
    # rather than letting someone enable them and hear nothing.
    smtp = _smtp_row()
    d["smtp_ready"] = bool(smtp.get("enabled") and (smtp.get("host") or "").strip())
    d["public_url"] = mailer.PUBLIC_URL or ""
    return d


@app.put("/api/projects/{project_id}/notify-config")
def put_notify_config(project_id: int, req: NotifyConfigReq, user=Depends(_proj_admin)):
    cur = _notify_cfg(project_id)
    if req.digest_cron:
        try:
            build_trigger(req.digest_cron)
        except Exception as exc:                   # noqa: BLE001
            raise HTTPException(400, f"Invalid digest schedule: {exc}")
    bad = mailer.invalid_recipients(req.recipients or "")
    if bad:
        raise HTTPException(400, "Not valid email addresses: " + ", ".join(bad[:5]))

    pick = lambda new, key: (cur.get(key) if new is None else (1 if new else 0))
    conn = get_db()
    conn.execute("""
        INSERT INTO notify_config (project_id, enabled, on_run_failed, on_scheduled_failed,
            on_run_passed, only_on_change, notify_triggerer, weekly_digest, digest_cron,
            recipients, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(project_id) DO UPDATE SET
            enabled=excluded.enabled, on_run_failed=excluded.on_run_failed,
            on_scheduled_failed=excluded.on_scheduled_failed,
            on_run_passed=excluded.on_run_passed, only_on_change=excluded.only_on_change,
            notify_triggerer=excluded.notify_triggerer, weekly_digest=excluded.weekly_digest,
            digest_cron=excluded.digest_cron, recipients=excluded.recipients,
            updated_at=excluded.updated_at""",
        (project_id, pick(req.enabled, "enabled"), pick(req.on_run_failed, "on_run_failed"),
         pick(req.on_scheduled_failed, "on_scheduled_failed"),
         pick(req.on_run_passed, "on_run_passed"), pick(req.only_on_change, "only_on_change"),
         pick(req.notify_triggerer, "notify_triggerer"), pick(req.weekly_digest, "weekly_digest"),
         req.digest_cron or cur.get("digest_cron") or "0 8 * * 1",
         cur.get("recipients") if req.recipients is None else req.recipients,
         datetime.now().isoformat(timespec="seconds")))
    conn.commit()
    conn.close()
    _reload_digests()
    audit(user, "notify.update", project_id=project_id, enabled=req.enabled,
          recipients=len(mailer.parse_recipients(req.recipients or "")),
          weekly_digest=req.weekly_digest)
    return get_notify_config(project_id, user)


@app.get("/api/admin/ai-config")
def get_ai_config(user=Depends(_require_sys_admin)):
    cfg = _get_ai_config()
    key = cfg.get("api_key") or ""
    return {
        "enabled":    cfg["enabled"],
        "api_base":   cfg["api_base"],
        "model":      cfg["model"],
        "verify_ssl": cfg["verify_ssl"],
        "has_key":    bool(key),
        "key_hint":   ("…" + key[-4:]) if len(key) > 4 else "",
    }


@app.put("/api/admin/ai-config")
def put_ai_config(req: AIConfigReq, user=Depends(_require_sys_admin)):
    conn = get_db()
    now = datetime.now().isoformat()
    # Upsert — a plain UPDATE silently no-ops if the id=1 row was never seeded
    conn.execute("INSERT OR IGNORE INTO ai_config (id, enabled) VALUES (1, 0)")
    if req.api_key:
        conn.execute(
            "UPDATE ai_config SET enabled=?, api_base=?, api_key=?, model=?, verify_ssl=?, updated_at=? WHERE id=1",
            (int(req.enabled), req.api_base, encrypt_token(req.api_key), req.model,
             int(req.verify_ssl), now),
        )
    else:
        conn.execute(
            "UPDATE ai_config SET enabled=?, api_base=?, model=?, verify_ssl=?, updated_at=? WHERE id=1",
            (int(req.enabled), req.api_base, req.model, int(req.verify_ssl), now),
        )
    conn.commit()
    saved = conn.execute("SELECT enabled, api_key FROM ai_config WHERE id=1").fetchone()
    conn.close()
    audit(user, "ai.config_update", enabled=req.enabled, model=req.model,
          api_base=req.api_base, api_key=bool(req.api_key))
    # Echo back what actually landed, so the UI never reports a save that didn't happen
    return {"ok": True,
            "enabled": bool(saved["enabled"]),
            "has_key": bool(saved["api_key"])}


@app.post("/api/admin/ai-config/test")
def test_ai_config(user=Depends(_require_sys_admin)):
    """Make a real minimal call so the admin gets a definitive answer, not a guess."""
    import requests
    cfg = _get_ai_config()
    if not cfg["api_key"]:
        raise HTTPException(400, "No API key saved")
    url = cfg["api_base"].rstrip("/") + "/chat/completions"
    _hush_insecure_warning(cfg["verify_ssl"])
    try:
        r = requests.post(
            url,
            headers={"Authorization": f"Bearer {cfg['api_key']}", "Content-Type": "application/json"},
            json={"model": cfg["model"],
                  "messages": [{"role": "user", "content": "Reply with the single word: ok"}],
                  "max_tokens": 8},
            timeout=30, verify=cfg["verify_ssl"],
        )
    except Exception as exc:                       # noqa: BLE001 — surfaced verbatim to the admin
        raise HTTPException(502, f"Cannot reach {url} — {type(exc).__name__}: {exc}")

    r.encoding = "utf-8"          # same ISO-8859-1 fallback applies to r.text
    if r.status_code != 200:
        raise HTTPException(502, f"HTTP {r.status_code} from provider: {r.text[:400]}")
    try:
        data = r.json()
        reply = data["choices"][0]["message"]["content"]
        model = data.get("model", cfg["model"])
    except (ValueError, KeyError, IndexError):
        raise HTTPException(502, f"Unexpected response shape: {r.text[:400]}")
    return {"ok": True, "model": model, "reply": (reply or "").strip()[:120]}


# ══════════════════════════════════════════════════════════════════
# DATA MANAGEMENT — bulk cleanup (project_admin / sys admin)
# ══════════════════════════════════════════════════════════════════

def _require_proj_admin(project_id: int, user: dict) -> None:
    _require_cap(project_id, user, "manage")


@app.get("/api/projects/{project_id}/data-summary")
def data_summary(project_id: int, user=Depends(_proj_viewer)):
    """Counts backing the Danger Zone, so the admin sees the blast radius first."""
    conn = get_db()
    q = lambda sql: conn.execute(sql, (project_id,)).fetchone()[0]      # noqa: E731
    out = {
        "test_cases": q("SELECT COUNT(*) FROM test_cases  WHERE project_id=?"),
        "suites":     q("SELECT COUNT(*) FROM test_groups WHERE project_id=?"),
        "runs":       q("SELECT COUNT(*) FROM test_runs   WHERE project_id=?"),
        "schedules":  q("SELECT COUNT(*) FROM schedules   WHERE project_id=?"),
    }
    conn.close()
    results_dir = _project_results(project_id)
    if results_dir.exists():
        out["result_dirs"] = sum(1 for p in results_dir.iterdir() if p.is_dir())
        out["result_bytes"] = sum(p.stat().st_size for p in results_dir.rglob("*") if p.is_file())
    else:
        out["result_dirs"] = 0
        out["result_bytes"] = 0
    return out


@app.get("/api/projects/{project_id}/tester-activity")
def tester_activity(project_id: int,
                    date_from: Optional[str] = None, date_to: Optional[str] = None,
                    user=Depends(_current_user)):
    """Per-tester execution activity for one project. project_admin only.

    Measures what the system actually records: runs triggered, tests executed,
    outcomes, machine execution time and active days. It does NOT measure hours
    worked — nothing in BRACE observes a person's working time.
    """
    _require_proj_admin(project_id, user)

    if not date_from and not date_to:                       # default: last 30 days
        date_to   = datetime.now().strftime("%Y-%m-%d")
        date_from = (datetime.now() - timedelta(days=29)).strftime("%Y-%m-%d")
    date_from = date_from or "0001-01-01"
    date_to   = date_to   or "9999-12-31"
    lo, hi = f"{date_from} 00:00:00", f"{date_to} 23:59:59"

    conn = get_db()
    runs = rows_to_list(conn.execute(
        """SELECT run_id, triggered_by, status, total, passed, failed, started_at, finished_at
           FROM test_runs
           WHERE project_id=? AND substr(replace(started_at,'T',' '),1,19) BETWEEN ? AND ?
           ORDER BY started_at, id""",
        (project_id, lo, hi)).fetchall())

    # Distinct test cases each person actually exercised
    cov = rows_to_list(conn.execute(
        """SELECT tr.triggered_by AS who, COUNT(DISTINCT tri.tc_code) AS n
           FROM test_run_items tri JOIN test_runs tr ON tri.run_id = tr.run_id
           WHERE tr.project_id=? AND substr(replace(tr.started_at,'T',' '),1,19) BETWEEN ? AND ?
           GROUP BY tr.triggered_by""",
        (project_id, lo, hi)).fetchall())
    conn.close()
    coverage = {c["who"]: c["n"] for c in cov}

    def _secs(a, b):
        if not a or not b:
            return None
        try:
            return max(0.0, (datetime.fromisoformat(b) - datetime.fromisoformat(a)).total_seconds())
        except ValueError:
            return None

    people: dict = {}
    daily:  dict = {}
    for r in runs:
        who = r["triggered_by"] or "(unknown)"
        day = (r["started_at"] or "")[:10]
        p = people.setdefault(who, {
            "user": who, "runs": 0, "tests": 0, "passed": 0, "failed": 0,
            "exec_seconds": 0.0, "days": set(), "first_seen": None, "last_seen": None,
            "cancelled": 0,
        })
        p["runs"]   += 1
        p["tests"]  += r["total"]  or 0
        p["passed"] += r["passed"] or 0
        p["failed"] += r["failed"] or 0
        if r["status"] == "cancelled":
            p["cancelled"] += 1
        d = _secs(r["started_at"], r["finished_at"])
        if d is not None:
            p["exec_seconds"] += d
        if day:
            p["days"].add(day)
            p["first_seen"] = min(p["first_seen"] or day, day)
            p["last_seen"]  = max(p["last_seen"]  or day, day)
            slot = daily.setdefault(day, {})
            slot[who] = slot.get(who, 0) + 1

    testers = []
    for p in people.values():
        active = len(p["days"]) or 1
        tests  = p["tests"]
        testers.append({
            "user":            p["user"],
            "runs":            p["runs"],
            "tests":           tests,
            "passed":          p["passed"],
            "failed":          p["failed"],
            "cancelled":       p["cancelled"],
            "pass_rate":       round(p["passed"] / tests * 100, 1) if tests else 0.0,
            "exec_seconds":    round(p["exec_seconds"], 1),
            "active_days":     len(p["days"]),
            "runs_per_day":    round(p["runs"] / active, 1),
            "tests_per_day":   round(tests / active, 1),
            "unique_tcs":      coverage.get(p["user"], 0),
            "first_seen":      p["first_seen"],
            "last_seen":       p["last_seen"],
        })
    testers.sort(key=lambda t: t["tests"], reverse=True)

    return {
        "date_from": date_from,
        "date_to":   date_to,
        "testers":   testers,
        "daily":     [{"date": d, "by_user": daily[d]} for d in sorted(daily)],
        "totals": {
            "testers":      len(testers),
            "runs":         sum(t["runs"]  for t in testers),
            "tests":        sum(t["tests"] for t in testers),
            "exec_seconds": round(sum(t["exec_seconds"] for t in testers), 1),
        },
    }


@app.delete("/api/projects/{project_id}/runs")
def purge_runs(project_id: int, keep_last: int = 0, user=Depends(_current_user)):
    """Delete run records and their result files. keep_last>0 retains the newest N."""
    _require_proj_admin(project_id, user)
    conn = get_db()
    if keep_last > 0:
        rows = conn.execute(
            """SELECT run_id FROM test_runs WHERE project_id=? AND run_id NOT IN
               (SELECT run_id FROM test_runs WHERE project_id=?
                ORDER BY started_at DESC, id DESC LIMIT ?)""",
            (project_id, project_id, keep_last),
        ).fetchall()
    else:
        rows = conn.execute("SELECT run_id FROM test_runs WHERE project_id=?",
                            (project_id,)).fetchall()
    # Never delete a run that's still executing
    run_ids = [r["run_id"] for r in rows if r["run_id"] not in _active_runs]
    skipped = len(rows) - len(run_ids)
    for rid in run_ids:
        conn.execute("DELETE FROM test_run_items WHERE run_id=?", (rid,))
        conn.execute("DELETE FROM test_runs     WHERE run_id=?", (rid,))
    conn.commit()
    conn.close()

    results_dir = _project_results(project_id)
    freed = 0
    for rid in run_ids:
        d = results_dir / rid
        if d.exists() and d.is_dir():
            freed += sum(p.stat().st_size for p in d.rglob("*") if p.is_file())
            shutil.rmtree(d, ignore_errors=True)
    audit(user, "data.purge_runs", project_id=project_id,
          deleted=len(run_ids), keep_last=keep_last, freed_bytes=freed)
    return {"deleted_runs": len(run_ids), "freed_bytes": freed, "skipped_active": skipped}


@app.delete("/api/projects/{project_id}/test-cases")
def purge_test_cases(project_id: int, user=Depends(_current_user)):
    _require_proj_admin(project_id, user)
    conn = get_db()
    ids = [r["id"] for r in conn.execute(
        "SELECT id FROM test_cases WHERE project_id=?", (project_id,)).fetchall()]
    if ids:
        marks = ",".join("?" * len(ids))
        conn.execute(f"UPDATE test_run_items SET test_case_id=NULL WHERE test_case_id IN ({marks})", ids)
        conn.execute(f"DELETE FROM group_test_cases WHERE test_case_id IN ({marks})", ids)
        conn.execute("DELETE FROM test_cases WHERE project_id=?", (project_id,))
    conn.commit()
    conn.close()
    audit(user, "data.purge_test_cases", project_id=project_id, deleted=len(ids))
    return {"deleted_test_cases": len(ids)}


@app.delete("/api/projects/{project_id}/groups")
def purge_suites(project_id: int, user=Depends(_current_user)):
    _require_proj_admin(project_id, user)
    conn = get_db()
    ids = [r["id"] for r in conn.execute(
        "SELECT id FROM test_groups WHERE project_id=?", (project_id,)).fetchall()]
    if ids:
        marks = ",".join("?" * len(ids))
        conn.execute(f"UPDATE test_runs SET group_id=NULL WHERE group_id IN ({marks})", ids)
        conn.execute(f"DELETE FROM schedules        WHERE group_id IN ({marks})", ids)
        conn.execute(f"DELETE FROM group_test_cases WHERE group_id IN ({marks})", ids)
        conn.execute("DELETE FROM test_groups WHERE project_id=?", (project_id,))
    conn.commit()
    conn.close()
    audit(user, "data.purge_suites", project_id=project_id, deleted=len(ids))
    return {"deleted_suites": len(ids)}


# ══════════════════════════════════════════════════════════════════
# SCHEDULES
# ══════════════════════════════════════════════════════════════════

class ScheduleCreate(BaseModel):
    group_id:  int
    cron_expr: str


class ScheduleUpdate(BaseModel):
    cron_expr: Optional[str]  = None
    enabled:   Optional[bool] = None


@app.get("/api/projects/{project_id}/schedules")
def list_schedules(project_id: int, user=Depends(_proj_viewer)):
    conn = get_db()
    rows = rows_to_list(conn.execute("""
        SELECT s.*, tg.name AS group_name
        FROM schedules s JOIN test_groups tg ON s.group_id = tg.id
        WHERE s.project_id=? ORDER BY s.created_at DESC
    """, (project_id,)).fetchall())
    for s in rows:
        s["enabled"]   = bool(s.get("enabled"))
        s["next_runs"] = next_run_times(s["cron_expr"], 3) if s["enabled"] else []
        last = conn.execute("""
            SELECT run_id, status, started_at, passed, failed FROM test_runs
            WHERE project_id=? AND group_id=? AND triggered_by='scheduler'
            ORDER BY started_at DESC, id DESC LIMIT 1
        """, (project_id, s["group_id"])).fetchone()
        s["last_run"] = dict(last) if last else None
    conn.close()
    return {"schedules": rows, "timezone": SCHEDULER_TZ}


@app.get("/api/admin/scheduler-jobs")
def scheduler_jobs(user=Depends(_require_sys_admin)):
    """What APScheduler actually holds right now — for diagnosing 'my schedule
    didn't fire' without reading pod logs."""
    return {"timezone": SCHEDULER_TZ, "jobs": scheduled_job_ids()}


@app.post("/api/projects/{project_id}/schedules/preview")
def preview_cron(project_id: int, req: dict, user=Depends(_proj_viewer)):
    """Validate a cron expression and show when it would next fire."""
    expr = (req.get("cron_expr") or "").strip()
    try:
        build_trigger(expr)
    except Exception as exc:                       # noqa: BLE001 — message goes to the user
        raise HTTPException(400, f"Invalid cron expression: {exc}")
    return {"valid": True, "timezone": SCHEDULER_TZ, "next_runs": next_run_times(expr, 5)}


@app.post("/api/projects/{project_id}/schedules")
def create_schedule(project_id: int, req: ScheduleCreate, user=Depends(_proj_tester)):
    try:
        build_trigger(req.cron_expr)               # real parse, not just a field count
    except Exception as exc:                       # noqa: BLE001
        raise HTTPException(400, f"Invalid cron expression: {exc}")
    conn = get_db()
    # Don't let a schedule point at another project's suite
    if not conn.execute("SELECT 1 FROM test_groups WHERE id=? AND project_id=?",
                        (req.group_id, project_id)).fetchone():
        conn.close()
        raise HTTPException(404, "Suite not found in this project")
    conn.execute("INSERT INTO schedules (project_id, group_id, cron_expr) VALUES (?,?,?)",
                 (project_id, req.group_id, req.cron_expr))
    conn.commit()
    sid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    _reload_all_jobs()
    audit(user, "schedule.create", project_id=project_id, target=str(sid),
          cron=req.cron_expr, group_id=req.group_id)
    return {"id": sid, "cron_expr": req.cron_expr}


@app.put("/api/schedules/{sid}")
def update_schedule(sid: int, req: ScheduleUpdate, user=Depends(_current_user)):
    conn = get_db()
    s = _owned_row(conn, "SELECT project_id FROM schedules WHERE id=?",
                   sid, user, "edit", "Schedule not found")
    if req.cron_expr:
        try:
            build_trigger(req.cron_expr)
        except Exception as exc:                   # noqa: BLE001
            conn.close()
            raise HTTPException(400, f"Invalid cron expression: {exc}")
        conn.execute("UPDATE schedules SET cron_expr=? WHERE id=?", (req.cron_expr, sid))
    if req.enabled is not None:
        conn.execute("UPDATE schedules SET enabled=? WHERE id=?", (1 if req.enabled else 0, sid))
    conn.commit()
    conn.close()
    _reload_all_jobs()
    audit(user, "schedule.update", project_id=s["project_id"], target=str(sid),
          cron=req.cron_expr, enabled=req.enabled)
    return {"ok": True}


@app.delete("/api/schedules/{sid}")
def delete_schedule(sid: int, user=Depends(_current_user)):
    conn = get_db()
    s = _owned_row(conn, "SELECT project_id, cron_expr FROM schedules WHERE id=?",
                   sid, user, "manage", "Schedule not found")
    conn.execute("DELETE FROM schedules WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    _reload_all_jobs()
    audit(user, "schedule.delete", project_id=s["project_id"], target=str(sid),
          cron=s["cron_expr"])
    return {"ok": True}


def _trigger_group_run(group_id: int):
    """Called by APScheduler from a worker thread.

    Must hand the work to the main event loop. The previous version spun up a
    throwaway loop, ran trigger_run in it (which only schedules the execution
    task) and then closed that loop — destroying the task before it ran, so
    scheduled runs were created as 'queued' and never executed.
    """
    conn = get_db()
    g = conn.execute("SELECT * FROM test_groups WHERE id=?", (group_id,)).fetchone()
    if not g:
        conn.close()
        return
    members = conn.execute("""
        SELECT tc.* FROM test_cases tc
        JOIN group_test_cases gtc ON tc.id = gtc.test_case_id
        WHERE gtc.group_id=? AND tc.project_id=? ORDER BY gtc.order_idx
    """, (group_id, g["project_id"])).fetchall()
    conn.close()
    tcs = rows_to_list(members)
    if not tcs:
        log.warning("Scheduled run for suite '%s' skipped — it has no test cases.", g["name"])
        return

    if _main_loop is None or _main_loop.is_closed():
        log.error("Scheduled run for '%s' skipped — no running event loop.", g["name"])
        return
    # Marshal onto the loop that owns the run executor and the slot semaphore
    _main_loop.call_soon_threadsafe(
        lambda: _start_run(g["project_id"], tcs, f"Scheduled: {g['name']}",
                           "scheduler", None, group_id=group_id))
    log.info("Scheduled run queued for suite '%s' (%d test cases).", g["name"], len(tcs))


# ══════════════════════════════════════════════════════════════════
# HOUSEKEEPING + AUDIT (admin)
# ══════════════════════════════════════════════════════════════════

@app.get("/api/admin/maintenance")
async def get_maintenance(refresh: bool = False, user=Depends(_require_sys_admin)):
    """Retention settings, current disk usage, and the last job's outcome.

    Disk size comes from the shared cache, so opening Administration is instant
    instead of waiting on a walk of the whole results volume. refresh=1 forces a
    fresh measurement for the operator who wants the live number.
    """
    disk = await asyncio.to_thread(maintenance.disk_usage, refresh)
    return {"config": maintenance.config(),
            "disk":   disk,
            "last":   maintenance.last_result(),
            "busy":   bool(_active_runs) or bool(_active_procs)}


@app.post("/api/admin/maintenance/run")
async def run_maintenance_now(dry_run: bool = True, user=Depends(_require_sys_admin)):
    """Run housekeeping on demand.

    Defaults to dry_run=True. Deleting run history is irreversible, so the
    caller has to ask for it explicitly — the UI shows what a dry run found and
    makes the operator confirm before the real thing.
    """
    busy = bool(_active_runs) or bool(_active_procs)
    # Blocking: rmtree over thousands of files, plus possibly a VACUUM.
    res = await asyncio.to_thread(maintenance.run_maintenance,
                                  set(_active_runs), not busy, dry_run)
    if not dry_run:
        audit(user, "maintenance.run_now",
              runs=res.get("runs", {}).get("runs", 0),
              orphan_dirs=res.get("orphans", {}).get("dirs", 0),
              freed_bytes=(res.get("runs", {}).get("freed_bytes", 0)
                           + res.get("orphans", {}).get("freed_bytes", 0)))
    return res


@app.get("/api/admin/audit")
def list_audit(username: str = "", action: str = "", project_id: int = 0,
               date_from: str = "", date_to: str = "",
               offset: int = 0, limit: int = 50,
               user=Depends(_require_sys_admin)):
    """Paged, filtered audit trail. Admin only — it names who did what."""
    limit  = max(1, min(500, limit))
    offset = max(0, offset)
    where, params = [], []
    if username:
        where.append("username=?");   params.append(username)
    if action:
        # 'run' matches run.trigger, run.cancel … — prefix, not substring, so a
        # filter cannot accidentally span unrelated action families.
        where.append("(action=? OR action LIKE ?)"); params += [action, f"{action}.%"]
    if project_id:
        where.append("project_id=?"); params.append(project_id)
    if date_from:
        where.append("substr(ts,1,10) >= ?"); params.append(date_from[:10])
    if date_to:
        where.append("substr(ts,1,10) <= ?"); params.append(date_to[:10])
    clause = (" WHERE " + " AND ".join(where)) if where else ""

    conn = get_db()
    try:
        total = conn.execute(f"SELECT COUNT(*) AS n FROM audit_log{clause}",
                             params).fetchone()["n"]
        rows = rows_to_list(conn.execute(
            f"SELECT * FROM audit_log{clause} ORDER BY id DESC LIMIT ? OFFSET ?",
            params + [limit, offset]).fetchall())
        # Distinct values for the filter dropdowns — cheap, both are indexed.
        users   = [r[0] for r in conn.execute(
            "SELECT DISTINCT username FROM audit_log ORDER BY username").fetchall() if r[0]]
        actions = [r[0] for r in conn.execute(
            "SELECT DISTINCT action FROM audit_log ORDER BY action").fetchall() if r[0]]
    finally:
        conn.close()
    return {"total": total, "offset": offset, "limit": limit, "entries": rows,
            "usernames": users, "actions": actions}


# ══════════════════════════════════════════════════════════════════
# HEALTH + STATIC
# ══════════════════════════════════════════════════════════════════

@app.get("/metrics", response_class=PlainTextResponse)
def metrics():
    """Prometheus text exposition. Hand-rolled to avoid a client dependency.

    Unauthenticated so an in-cluster scraper can reach it; it exposes only
    aggregate counters, never test content, project names or user identities.
    """
    m, out = _metrics, []

    def emit(name, mtype, help_text, value, labels=""):
        out.append(f"# HELP {name} {help_text}")
        out.append(f"# TYPE {name} {mtype}")
        out.append(f"{name}{labels} {value}")

    active = sum(1 for r in _active_runs.values() if r["status"] == "running")
    queued = sum(1 for r in _active_runs.values() if r["status"] == "queued")

    emit("brace_up", "gauge", "1 when the controller is serving.", 1)
    emit("brace_uptime_seconds", "gauge", "Seconds since process start.",
         round((datetime.now() - m["started_at"]).total_seconds(), 1))
    emit("brace_runs_active", "gauge", "Runs currently executing.", active)
    emit("brace_runs_queued", "gauge", "Runs waiting for an execution slot.", queued)
    emit("brace_run_slots_total", "gauge", "Configured concurrent run limit.",
         MAX_CONCURRENT_RUNS)
    emit("brace_run_slots_available", "gauge", "Free execution slots.",
         max(0, MAX_CONCURRENT_RUNS - active))
    # Test cases now execute in parallel within a run, so runs no longer track
    # browser count. These are the gauges that reflect actual pod load.
    emit("brace_tests_running", "gauge", "Robot processes (browsers) executing now.",
         len(_active_procs))
    emit("brace_test_slots_total", "gauge", "Configured concurrent robot-process limit.",
         MAX_CONCURRENT_TESTS)
    emit("brace_test_slots_available", "gauge", "Free robot-process slots.",
         max(0, MAX_CONCURRENT_TESTS - len(_active_procs)))

    out.append("# HELP brace_runs_total Runs by final state since start.")
    out.append("# TYPE brace_runs_total counter")
    for state, key in (("started", "runs_started"), ("passed", "runs_passed"),
                       ("failed", "runs_failed"), ("cancelled", "runs_cancelled")):
        out.append(f'brace_runs_total{{state="{state}"}} {m[key]}')

    out.append("# HELP brace_tests_total Test cases executed by outcome since start.")
    out.append("# TYPE brace_tests_total counter")
    for state, key in (("passed", "tests_passed"), ("failed", "tests_failed"),
                       ("timeout", "tests_timeout")):
        out.append(f'brace_tests_total{{outcome="{state}"}} {m[key]}')

    emit("brace_test_duration_seconds_sum", "counter",
         "Total test-case execution seconds.", round(m["test_seconds"], 1))
    emit("brace_test_duration_seconds_count", "counter",
         "Test cases contributing to the duration sum.", m["test_count"])

    # Disk is the failure mode most likely to take the pod down quietly.
    # Read from the cache: measuring it costs a full recursive walk of the
    # results volume (~7.5s at 384 MB, and it only grows). Doing that per scrape
    # meant the walk ran almost continuously and would eventually outlive the
    # scrape timeout — the monitoring would fail before the disk did.
    _disk = maintenance.results_bytes()
    emit("brace_results_disk_bytes", "gauge",
         "Bytes consumed by run results on disk. Sampled, not live — see "
         "brace_results_disk_age_seconds.", _disk["bytes"])
    emit("brace_results_disk_age_seconds", "gauge",
         "Age of the results-disk sample. Compare with BRACE_DISK_CACHE_TTL.",
         _disk["age_sec"])
    try:
        db_bytes = DB_PATH.stat().st_size
    except OSError:
        db_bytes = -1
    emit("brace_database_bytes", "gauge", "Size of the SQLite database file.", db_bytes)

    emit("brace_retention_days", "gauge",
         "Configured run retention in days; 0 means keep forever.",
         maintenance.RETENTION_DAYS)
    _last = maintenance.last_result()
    if _last:
        emit("brace_maintenance_last_runs_purged", "gauge",
             "Runs deleted by the most recent housekeeping job.",
             _last.get("runs", {}).get("runs", 0))
        emit("brace_maintenance_last_freed_bytes", "gauge",
             "Bytes reclaimed by the most recent housekeeping job.",
             _last.get("runs", {}).get("freed_bytes", 0)
             + _last.get("orphans", {}).get("freed_bytes", 0))

    try:
        conn = get_db()
        for name, sql in (("projects",   "SELECT COUNT(*) FROM projects"),
                          ("test_cases", "SELECT COUNT(*) FROM test_cases"),
                          ("runs",       "SELECT COUNT(*) FROM test_runs"),
                          # The table that actually drives DB growth — runs x cases.
                          ("run_items",  "SELECT COUNT(*) FROM test_run_items")):
            emit(f"brace_{name}_count", "gauge", f"Rows in {name}.",
                 conn.execute(sql).fetchone()[0])
        conn.close()
    except Exception:                                  # noqa: BLE001 — metrics must not 500
        pass

    return "\n".join(out) + "\n"


@app.get("/health")
def health():
    """Liveness: is this process still serving? Nothing else.

    This deliberately does NOT touch the database. It used to, and under load
    that was actively harmful: with several tests executing, SQLite gets
    contended, get_db() can raise 'database is locked', /health returns 500,
    and the kubelet kills the pod — destroying every run in flight over a
    condition that would have cleared itself in a second.

    A liveness probe should only answer "is this wedged beyond recovery?".
    Anything that can fail transiently belongs in readiness, below.

    Public (probes hit it unauthenticated), so it must not expose tenant data.
    """
    return {
        "status":    "ok",
        "version":   APP_VERSION,
        "bss_env":   BSS_ENV,
        "pod_name":  POD_NAME,
        "image_tag": IMAGE_TAG,
    }


@app.get("/health/ready")
def health_ready(response: Response):
    """Readiness: can this pod actually serve requests right now?

    Checks the database, but reports 503 rather than raising — readiness only
    removes the pod from the Service, it never restarts it. With replicas: 1
    that briefly returns 503 to the route, which is the correct, recoverable
    behaviour when the DB is momentarily busy.
    """
    try:
        conn = get_db()
        conn.execute("SELECT 1").fetchone()
        conn.close()
    except Exception as exc:               # noqa: BLE001 — reported, not raised
        log.warning("Readiness check failed: %s", exc)
        response.status_code = 503
        return {"status": "degraded", "detail": str(exc)[:200]}
    return {"status": "ok", "tests_running": len(_active_procs)}


@app.get("/api/run-config")
def run_config(user=Depends(_current_user)):
    """Execution limits the run dialog needs to offer a sensible parallel box."""
    return {"max_parallel": MAX_CONCURRENT_TESTS,
            "default_parallel": RUN_PARALLEL_DEFAULT}


@app.get("/api/health-detail")
def health_detail(user=Depends(_require_sys_admin)):
    conn = get_db()
    projects = conn.execute("SELECT COUNT(*) FROM projects").fetchone()[0]
    users    = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    conn.close()
    return {"status": "ok", "version": APP_VERSION, "bss_env": BSS_ENV,
            "pod_name": POD_NAME, "image_tag": IMAGE_TAG,
            "projects": projects, "users": users,
            "max_concurrent_runs":  MAX_CONCURRENT_RUNS,
            "max_concurrent_tests": MAX_CONCURRENT_TESTS,
            "run_parallel_default": RUN_PARALLEL_DEFAULT,
            "tests_running":        len(_active_procs),
            "test_timeout_sec":     TEST_TIMEOUT_SEC}


# Resolves to /opt/rf/controller/static in the image; derived rather than
# hardcoded so the app also runs outside the container (tests, local dev).
STATIC_DIR = Path(os.getenv("STATIC_DIR") or (Path(__file__).resolve().parent / "static"))
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/", response_class=HTMLResponse)
@app.get("/{path:path}", response_class=HTMLResponse, include_in_schema=False)
def serve_spa(path: str = ""):
    html = STATIC_DIR / "index.html"
    if html.exists():
        return HTMLResponse(html.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>BRACE — UI not found</h1>")
